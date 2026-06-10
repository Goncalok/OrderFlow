from __future__ import annotations

import json
import os
import re
import tempfile
import unicodedata
import uuid
from dataclasses import replace
from datetime import datetime
from email import policy
from email.parser import BytesParser
from pathlib import Path
from typing import Any

from email_order_app.excel_writer import (
    fill_cmr_template,
    fill_colruyt_cmr_template,
    fill_denemark_cmr_template,
    fill_edeka_laatzen_cmr_template,
    fill_edeka_mochmuhl_cmr_template,
    fill_havi_nl_cmr_template,
    fill_havi_be_cmr_template,
    fill_havi_de_cmr_template,
    fill_havi_duisburg_cmr_template,
    fill_havi_wunstorf_cmr_template,
    fill_havi_neu_wulmstorf_cmr_template,
    fill_heeren_cmr_template,
    fill_nettomd_cmr_template,
    fill_rewe_penny_cmr_template,
    fill_hanos_cmr_template,
    fill_template,
)
from email_order_app.leverschema_writer import export_leverschema_workbook
from email_order_app.models import CustomerOrder, OrderItem, ParsedOrderEmail
from email_order_app.parser import parse_order_email
from email_order_app.special_netto_md import NettoMdEmail, parse_netto_md_email, write_netto_md_excel
from email_order_app.special_havi_uien import HaviUienOrder, parse_havi_uien_email, write_havi_uien_excel


BASE_DIR = Path(__file__).resolve().parent.parent
TEMPLATE_PATH = BASE_DIR / "Piking file.xlsx"
REWE_PENNY_TEMPLATE_PATH = BASE_DIR / "ORDER REWE PENNY Print In Color.xlsx"
LEVERSCHEMA_TEMPLATE_PATH = BASE_DIR / "Leverschema OOH.xlsm"
NETTO_MD_TEMPLATE_PATH = BASE_DIR / "Order picking NETTO blanco.xlsx"
CMR_TEMPLATE_PATH = BASE_DIR / "CMR" / "FIF CMR DOCUMENT.xlsm"
KDC_CMR_TEMPLATE_PATH = BASE_DIR / "CMR" / "KDC CMR DOCUMENT.xlsm"
COLRUYT_CMR_TEMPLATE_PATH = BASE_DIR / "CMR" / "CMR Colruyt.xlsm"
DENEMARK_CMR_TEMPLATE_PATH = BASE_DIR / "CMR" / "CMR Denemarken.xls"
EDEKA_LAATZEN_CMR_TEMPLATE_PATH = BASE_DIR / "CMR" / "CMR EDEKA_DC_LAATZEN.xlsm"
EDEKA_MOCHMUHL_CMR_TEMPLATE_PATH = BASE_DIR / "CMR" / "CMR EDEKA DC MOCKMUHL.xlsm"
GLOBUS_CMR_TEMPLATE_PATH = BASE_DIR / "CMR" / "CMR Globus.xlsm"
HAVI_NL_CMR_TEMPLATE_PATH = BASE_DIR / "CMR" / "CMR NL (Havi).xlsm"
HAVI_BE_CMR_TEMPLATE_PATH = BASE_DIR / "CMR" / "CMR BEL.xlsm"
HAVI_DE_CMR_TEMPLATE_PATH = BASE_DIR / "CMR" / "CMR GERMANY.xlsm"
HAVI_DUISBURG_CMR_TEMPLATE_PATH = BASE_DIR / "CMR" / "CMR Havi Logistics (DC Duisburg).xlsm"
HAVI_WUNSTORF_CMR_TEMPLATE_PATH = BASE_DIR / "CMR" / "CMR Havi Logistics (DC Wunstorf).xlsm"
HAVI_NEU_WULMSTORF_CMR_TEMPLATE_PATH = BASE_DIR / "CMR" / "CMR Havi Logistics (DC Neu Wulmstorf).xlsm"
HEEREN_CMR_TEMPLATE_PATH = BASE_DIR / "CMR" / "Herren tbv Foodservice cmr.xlsm"
NETTOMD_KERPEN_CMR_TEMPLATE_PATH = BASE_DIR / "CMR" / "KERPEN.xls"
NETTOMD_HODENHAGEN_CMR_TEMPLATE_PATH = BASE_DIR / "CMR" / "HODENHAGEN.xls"
NETTOMD_HENSTEDT_ULZBURG_CMR_TEMPLATE_PATH = BASE_DIR / "CMR" / "HENSTEDT-UILZBURG.xls"
NETTOMD_HAMM_CMR_TEMPLATE_PATH = BASE_DIR / "CMR" / "HAMM.xls"
NETTOMD_GANDERKESEE_CMR_TEMPLATE_PATH = BASE_DIR / "CMR" / "GANDERKESEE.xls"
NETTOMD_BOTTROP_CMR_TEMPLATE_PATH = BASE_DIR / "CMR" / "BOTTROP.xls"
NETTOMD_KREFELD_CMR_TEMPLATE_PATH = BASE_DIR / "CMR" / "KREFELD.xls"
REWE_PENNY_CMR_TEMPLATE_PATH = BASE_DIR / "CMR" / "REWE-PENNY CMR.xlsm"
HANOS_CMR_TEMPLATE_PATH = BASE_DIR / "CMR" / "Hanos CMR.xlsm"
LAADSCHEMA_TEMPLATE_PATH = BASE_DIR / "Laadschema Outbound (Blanco).xlsx"
WORK_DIR = Path(tempfile.gettempdir()) / "email-order-web-app"


def resolve_netto_md_template_path() -> Path:
    """Prefer NETTO_MD_TEMPLATE_PATH env (absolute path on server); else project root template."""
    env = (os.environ.get("NETTO_MD_TEMPLATE_PATH") or "").strip()
    if env:
        candidate = Path(env)
        if candidate.is_file():
            return candidate
    primary = BASE_DIR / "Order picking NETTO blanco.xlsx"
    if primary.is_file():
        return primary
    alt = BASE_DIR / "Order picking NETTO blanco.XLSX"
    return alt if alt.is_file() else primary


def _is_definitely_netto_md_email(email_path: Path) -> bool:
    """
    Strict NettoMD detection to prevent interference with other clients.
    Only returns True if we're absolutely certain this is a NettoMD email.
    """
    try:
        from email import policy
        from email.parser import BytesParser
        
        with email_path.open("rb") as handle:
            message = BytesParser(policy=policy.default).parse(handle)
        
        # Check subject first - NettoMD emails have very specific subject patterns
        subject = str(message.get("subject", "")).strip().lower()
        if not ("netto md" in subject or "netto" in subject):
            return False
        
        # Extract email content
        from email_order_app.special_netto_md import _extract_plain_text
        body = _extract_plain_text(message).replace("\r", "")
        lines = [line.strip() for line in body.split("\n") if line.strip()]
        
        # Must have NettoMD-specific indicators
        content_text = " ".join(lines[:50]).lower()
        
        # STRICT requirements - ALL must be present for NettoMD
        required_indicators = [
            "netto lager",  # NettoMD uses "Netto lager [Location]"
            "netto md",     # NettoMD specific product names
        ]
        
        # Check if ALL required indicators are present
        has_all_indicators = all(
            indicator in content_text 
            for indicator in required_indicators
        )
        
        if not has_all_indicators:
            return False
        
        # Additional verification: check for NettoMD location patterns
        netto_locations = [
            "bottrop", "ganderkesee", "hamm", "henstedt", "hodenhagen", 
            "kerpen", "krefeld"
        ]
        
        has_netto_location = any(
            location in content_text 
            for location in netto_locations
        )
        
        return has_netto_location
        
    except Exception:
        return False


def parse_uploaded_email(file_name: str, file_bytes: bytes) -> dict[str, Any]:
    WORK_DIR.mkdir(parents=True, exist_ok=True)
    temp_path = WORK_DIR / f"{uuid.uuid4().hex}_{Path(file_name).name}"
    temp_path.write_bytes(file_bytes)
    try:
        # First check for special Havi Uien format
        special = parse_havi_uien_email(temp_path)
        if special is not None:
            return {"mode": "special", "preview": serialize_special(special)}

        # Check for REWE/Penny emails BEFORE NettoMD to avoid false detection
        # Quick check: if subject or filename contains REWE/Penny, use standard parser
        subject_check = file_name.lower() if file_name else ""
        if "rewe" in subject_check or "penny" in subject_check:
            parsed = repair_parsed_texts(parse_order_email(temp_path))
            return {"mode": "standard", "preview": serialize_standard(parsed)}
        
        # Also check email subject from the actual email
        try:
            from email import policy
            from email.parser import BytesParser
            with temp_path.open("rb") as handle:
                message = BytesParser(policy=policy.default).parse(handle)
            subject = str(message.get("subject", "")).strip().lower()
            if "rewe" in subject or "penny" in subject:
                parsed = repair_parsed_texts(parse_order_email(temp_path))
                return {"mode": "standard", "preview": serialize_standard(parsed)}
        except:
            pass  # If email parsing fails, continue with normal flow

        # Check NettoMD ONLY if very specific indicators are present
        # This prevents NettoMD from interfering with other clients
        if _is_definitely_netto_md_email(temp_path):
            netto_md = parse_netto_md_email(temp_path)
            if netto_md is not None:
                return {"mode": "netto_md", "preview": serialize_netto_md(netto_md)}

        # Finally, use standard parser for all other emails
        parsed = repair_parsed_texts(parse_order_email(temp_path))
        return {"mode": "standard", "preview": serialize_standard(parsed)}
    finally:
        if temp_path.exists():
            temp_path.unlink()


def export_uploaded_email(
    file_name: str | None,
    file_bytes: bytes | None,
    export_type: str,
    order_index: int = 0,
    quantity_overrides: dict[str, Any] | None = None,
    export_sheet: str | None = None,
    selected_client: str | None = None,
    cmr_references: list[str] | None = None,
    cmr_pallet_places: str | None = None,
    dc_name: str | None = None,
    pakbon_items: list[dict[str, Any]] | None = None,
    goederen_total: float | None = None,
    merge_order_indexes: list[int] | None = None,
    havi_uien_article: str | None = None,
    havi_uien_description: str | None = None,
) -> tuple[str, bytes]:
    WORK_DIR.mkdir(parents=True, exist_ok=True)
    
    # Handle Denemark CMR export without email
    client_lower = (selected_client or "").strip().lower()
    if export_type == "print_cmr" and ("denemark" in client_lower or "denmark" in client_lower):
        if not file_name or not file_bytes:
            output_name = "CMR Denemark.xls"
            output_path = WORK_DIR / f"{uuid.uuid4().hex}_{output_name}"
            try:
                # Create empty parsed email for template
                from datetime import datetime
                parsed = ParsedOrderEmail(
                    source_file="",
                    subject="",
                    sender="",
                    received_at=datetime.now(),
                    delivery_date_to_dc="",
                    leaving_date_venlo="",
                    leaving_time_venlo="",
                    orders=[],
                )
                order = CustomerOrder(
                    customer_dc="Denemark",
                    fatrans_dc="Denemark",
                    reference="",
                    items=[],
                    slotboeking_id=None,
                )
                fill_denemark_cmr_template(DENEMARK_CMR_TEMPLATE_PATH, output_path, parsed, order)

                output_bytes = output_path.read_bytes()

                return output_name, output_bytes
            finally:
                if output_path.exists():
                    output_path.unlink()
    
    # Handle Edeka Laatzen CMR export (with or without email)
    if export_type == "print_cmr" and "edeka laatzen" in client_lower:
        output_name = "CMR Edeka Laatzen.xlsm"
        output_path = WORK_DIR / f"{uuid.uuid4().hex}_{output_name}"
        try:
            if not file_name or not file_bytes:
                # No email - create empty data
                from datetime import datetime
                parsed = ParsedOrderEmail(
                    source_file="",
                    subject="",
                    sender="",
                    received_at=datetime.now(),
                    delivery_date_to_dc="",
                    leaving_date_venlo="",
                    leaving_time_venlo="",
                    orders=[],
                )
                order = CustomerOrder(
                    customer_dc="Edeka Laatzen",
                    fatrans_dc="Edeka",
                    reference="",
                    items=[],
                    slotboeking_id=None,
                )
            else:
                # With email - parse it
                email_path = WORK_DIR / f"{uuid.uuid4().hex}_{Path(file_name).name}"
                email_path.write_bytes(file_bytes)
                try:
                    parsed = parse_order_email(email_path)
                    if quantity_overrides:
                        _apply_standard_quantity_overrides(parsed, quantity_overrides)
                    if export_type == "merge":
                        order = merge_orders(parsed.orders)
                    else:
                        order = parsed.orders[order_index]
                    order = _apply_client_export_rules(order, selected_client)
                finally:
                    if email_path.exists():
                        email_path.unlink()
            
            fill_edeka_laatzen_cmr_template(EDEKA_LAATZEN_CMR_TEMPLATE_PATH, output_path, parsed, order)
            output_bytes = output_path.read_bytes()
            return output_name, output_bytes
        finally:
            if output_path.exists():
                output_path.unlink()
    
    # Handle Edeka Mochmuhl CMR export (with or without email)
    if export_type == "print_cmr" and "edeka mochmuhl" in client_lower:
        output_name = "CMR Edeka Mochmuhl.xlsm"
        output_path = WORK_DIR / f"{uuid.uuid4().hex}_{output_name}"
        try:
            if not file_name or not file_bytes:
                # No email - create empty data
                from datetime import datetime
                parsed = ParsedOrderEmail(
                    source_file="",
                    subject="",
                    sender="",
                    received_at=datetime.now(),
                    delivery_date_to_dc="",
                    leaving_date_venlo="",
                    leaving_time_venlo="",
                    orders=[],
                )
                order = CustomerOrder(
                    customer_dc="Edeka Mochmuhl",
                    fatrans_dc="Edeka",
                    reference="",
                    items=[],
                    slotboeking_id=None,
                )
            else:
                # With email - parse it
                email_path = WORK_DIR / f"{uuid.uuid4().hex}_{Path(file_name).name}"
                email_path.write_bytes(file_bytes)
                try:
                    parsed = parse_order_email(email_path)
                    if quantity_overrides:
                        _apply_standard_quantity_overrides(parsed, quantity_overrides)
                    if export_type == "merge":
                        order = merge_orders(parsed.orders)
                    else:
                        order = parsed.orders[order_index]
                    order = _apply_client_export_rules(order, selected_client)
                finally:
                    if email_path.exists():
                        email_path.unlink()
            
            fill_edeka_mochmuhl_cmr_template(EDEKA_MOCHMUHL_CMR_TEMPLATE_PATH, output_path, parsed, order)

            
            output_bytes = output_path.read_bytes()

            
            return output_name, output_bytes
        finally:
            if output_path.exists():
                output_path.unlink()
    
    # Handle Globus CMR export (with or without email) - same structure as Edeka
    if export_type == "print_cmr" and "globus" in client_lower:
        output_name = "CMR Globus.xlsm"
        output_path = WORK_DIR / f"{uuid.uuid4().hex}_{output_name}"
        try:
            if not file_name or not file_bytes:
                # No email - create empty data
                from datetime import datetime
                parsed = ParsedOrderEmail(
                    source_file="",
                    subject="",
                    sender="",
                    received_at=datetime.now(),
                    delivery_date_to_dc="",
                    leaving_date_venlo="",
                    leaving_time_venlo="",
                    orders=[],
                )
                order = CustomerOrder(
                    customer_dc="Globus",
                    fatrans_dc="Globus",
                    reference="",
                    items=[],
                    slotboeking_id=None,
                )
            else:
                # With email - parse it
                email_path = WORK_DIR / f"{uuid.uuid4().hex}_{Path(file_name).name}"
                email_path.write_bytes(file_bytes)
                try:
                    parsed = parse_order_email(email_path)
                    if quantity_overrides:
                        _apply_standard_quantity_overrides(parsed, quantity_overrides)
                    if export_type == "merge":
                        order = merge_orders(parsed.orders)
                    else:
                        order = parsed.orders[order_index]
                    order = _apply_client_export_rules(order, selected_client)
                finally:
                    if email_path.exists():
                        email_path.unlink()
            
            # Reuse Edeka function since structure is the same (A26:AG27 merged cells)
            fill_edeka_mochmuhl_cmr_template(GLOBUS_CMR_TEMPLATE_PATH, output_path, parsed, order)

            output_bytes = output_path.read_bytes()

            return output_name, output_bytes
        finally:
            if output_path.exists():
                output_path.unlink()

    # Handle Havi 3 DCs CMR export (Duisburg, Wunstorf, Neu Wulmstorf)
    if export_type == "print_cmr" and "havi" in client_lower and "duisburg" in client_lower and dc_name:
        dc_lower = dc_name.lower()
        if "duisburg" in dc_lower:
            template_path = HAVI_DUISBURG_CMR_TEMPLATE_PATH
            output_name = "CMR Havi Duisburg.xlsm"
            fill_function = fill_havi_duisburg_cmr_template
        elif "wunstorf" in dc_lower:
            template_path = HAVI_WUNSTORF_CMR_TEMPLATE_PATH
            output_name = "CMR Havi Wunstorf.xlsm"
            fill_function = fill_havi_wunstorf_cmr_template
        elif "neu wulmstorf" in dc_lower:
            template_path = HAVI_NEU_WULMSTORF_CMR_TEMPLATE_PATH
            output_name = "CMR Havi Neu Wulmstorf.xlsm"
            fill_function = fill_havi_neu_wulmstorf_cmr_template
        else:
            # Default to Duisburg if DC not specified
            template_path = HAVI_DUISBURG_CMR_TEMPLATE_PATH
            output_name = "CMR Havi Duisburg.xlsm"
            fill_function = fill_havi_duisburg_cmr_template
        
        output_path = WORK_DIR / f"{uuid.uuid4().hex}_{output_name}"
        try:
            if not file_name or not file_bytes:
                # No email - create empty data
                from datetime import datetime
                parsed = ParsedOrderEmail(
                    source_file="",
                    subject="",
                    sender="",
                    received_at=datetime.now(),
                    delivery_date_to_dc="",
                    leaving_date_venlo="",
                    leaving_time_venlo="",
                    orders=[],
                )
                order = CustomerOrder(
                    customer_dc=f"Havi {dc_name}",
                    fatrans_dc=f"Havi {dc_name}",
                    reference="",
                    items=[],
                    slotboeking_id=None,
                )
            else:
                # With email - parse it
                email_path = WORK_DIR / f"{uuid.uuid4().hex}_{Path(file_name).name}"
                email_path.write_bytes(file_bytes)
                try:
                    parsed = parse_order_email(email_path)
                    if quantity_overrides:
                        _apply_standard_quantity_overrides(parsed, quantity_overrides)
                    if export_type == "merge":
                        order = merge_orders(parsed.orders)
                    else:
                        order = parsed.orders[order_index]
                    order = _apply_client_export_rules(order, selected_client)
                finally:
                    if email_path.exists():
                        email_path.unlink()
            
            fill_function(template_path, output_path, parsed, order, references=cmr_references)

            
            output_bytes = output_path.read_bytes()

            
            return output_name, output_bytes
        finally:
            if output_path.exists():
                output_path.unlink()
    
    # Handle Havi DE CMR export (with or without email)
    if export_type == "print_cmr" and ("havi de" in client_lower or "havi de saturday" in client_lower):
        output_name = "CMR Havi DE.xlsm"
        output_path = WORK_DIR / f"{uuid.uuid4().hex}_{output_name}"
        try:
            if not file_name or not file_bytes:
                # No email - create empty data
                from datetime import datetime
                parsed = ParsedOrderEmail(
                    source_file="",
                    subject="",
                    sender="",
                    received_at=datetime.now(),
                    delivery_date_to_dc="",
                    leaving_date_venlo="",
                    leaving_time_venlo="",
                    orders=[],
                )
                order = CustomerOrder(
                    customer_dc="Havi DE",
                    fatrans_dc="Havi DE",
                    reference="",
                    items=[],
                    slotboeking_id=None,
                )
            else:
                # With email - parse it
                email_path = WORK_DIR / f"{uuid.uuid4().hex}_{Path(file_name).name}"
                email_path.write_bytes(file_bytes)
                try:
                    parsed = parse_order_email(email_path)
                    if quantity_overrides:
                        _apply_standard_quantity_overrides(parsed, quantity_overrides)
                    if export_type == "merge":
                        order = merge_orders(parsed.orders)
                    else:
                        order = parsed.orders[order_index]
                    order = _apply_client_export_rules(order, selected_client)
                finally:
                    if email_path.exists():
                        email_path.unlink()
            
            fill_havi_de_cmr_template(HAVI_DE_CMR_TEMPLATE_PATH, output_path, parsed, order)

            
            output_bytes = output_path.read_bytes()

            
            return output_name, output_bytes
        finally:
            if output_path.exists():
                output_path.unlink()
    
    # Handle Havi BE CMR export (with or without email)
    if export_type == "print_cmr" and "havi be" in client_lower:
        output_name = "CMR Havi BE.xlsm"
        output_path = WORK_DIR / f"{uuid.uuid4().hex}_{output_name}"
        try:
            if not file_name or not file_bytes:
                # No email - create empty data
                from datetime import datetime
                parsed = ParsedOrderEmail(
                    source_file="",
                    subject="",
                    sender="",
                    received_at=datetime.now(),
                    delivery_date_to_dc="",
                    leaving_date_venlo="",
                    leaving_time_venlo="",
                    orders=[],
                )
                order = CustomerOrder(
                    customer_dc="Havi BE",
                    fatrans_dc="Havi BE",
                    reference="",
                    items=[],
                    slotboeking_id=None,
                )
            else:
                # With email - parse it
                email_path = WORK_DIR / f"{uuid.uuid4().hex}_{Path(file_name).name}"
                email_path.write_bytes(file_bytes)
                try:
                    parsed = parse_order_email(email_path)
                    if quantity_overrides:
                        _apply_standard_quantity_overrides(parsed, quantity_overrides)
                    if export_type == "merge":
                        order = merge_orders(parsed.orders)
                    else:
                        order = parsed.orders[order_index]
                    order = _apply_client_export_rules(order, selected_client)
                finally:
                    if email_path.exists():
                        email_path.unlink()
            
            fill_havi_be_cmr_template(HAVI_BE_CMR_TEMPLATE_PATH, output_path, parsed, order, references=cmr_references)

            
            output_bytes = output_path.read_bytes()

            
            return output_name, output_bytes
        finally:
            if output_path.exists():
                output_path.unlink()
    
    # Handle Havi NL CMR export (with or without email)
    if export_type == "print_cmr" and "havi nl" in client_lower:
        output_name = "CMR Havi NL.xlsm"
        output_path = WORK_DIR / f"{uuid.uuid4().hex}_{output_name}"
        try:
            if not file_name or not file_bytes:
                # No email - create empty data
                from datetime import datetime
                parsed = ParsedOrderEmail(
                    source_file="",
                    subject="",
                    sender="",
                    received_at=datetime.now(),
                    delivery_date_to_dc="",
                    leaving_date_venlo="",
                    leaving_time_venlo="",
                    orders=[],
                )
                order = CustomerOrder(
                    customer_dc="Havi NL",
                    fatrans_dc="Havi NL",
                    reference="",
                    items=[],
                    slotboeking_id=None,
                )
            else:
                # With email - parse it
                email_path = WORK_DIR / f"{uuid.uuid4().hex}_{Path(file_name).name}"
                email_path.write_bytes(file_bytes)
                try:
                    parsed = parse_order_email(email_path)
                    if quantity_overrides:
                        _apply_standard_quantity_overrides(parsed, quantity_overrides)
                    if export_type == "merge":
                        order = merge_orders(parsed.orders)
                    else:
                        order = parsed.orders[order_index]
                    order = _apply_client_export_rules(order, selected_client)
                finally:
                    if email_path.exists():
                        email_path.unlink()
            
            fill_havi_nl_cmr_template(HAVI_NL_CMR_TEMPLATE_PATH, output_path, parsed, order, references=cmr_references)

            
            output_bytes = output_path.read_bytes()

            
            return output_name, output_bytes
        finally:
            if output_path.exists():
                output_path.unlink()

    # Handle Heeren CMR export (with or without email)
    if export_type == "print_cmr" and "heeren" in client_lower:
        output_name = "CMR Heeren.xlsm"
        output_path = WORK_DIR / f"{uuid.uuid4().hex}_{output_name}"
        try:
            if not file_name or not file_bytes:
                # No email - create empty data
                from datetime import datetime
                parsed = ParsedOrderEmail(
                    source_file="",
                    subject="",
                    sender="",
                    received_at=datetime.now(),
                    delivery_date_to_dc="",
                    leaving_date_venlo="",
                    leaving_time_venlo="",
                    orders=[],
                )
                order = CustomerOrder(
                    customer_dc="Heeren",
                    fatrans_dc="Heeren",
                    reference="",
                    items=[],
                    slotboeking_id=None,
                )
            else:
                # With email - parse it
                email_path = WORK_DIR / f"{uuid.uuid4().hex}_{Path(file_name).name}"
                email_path.write_bytes(file_bytes)
                try:
                    parsed = parse_order_email(email_path)
                    if quantity_overrides:
                        _apply_standard_quantity_overrides(parsed, quantity_overrides)
                    if export_type == "merge":
                        order = merge_orders(parsed.orders)
                    else:
                        order = parsed.orders[order_index]
                    order = _apply_client_export_rules(order, selected_client)
                finally:
                    if email_path.exists():
                        email_path.unlink()
            
            fill_heeren_cmr_template(HEEREN_CMR_TEMPLATE_PATH, output_path, parsed, order, references=cmr_references)

            
            output_bytes = output_path.read_bytes()

            
            return output_name, output_bytes
        finally:
            if output_path.exists():
                output_path.unlink()

    # Handle NettoMD CMR export (7 locations, always active)
    if export_type == "print_cmr" and "nettomd" in client_lower:
        if not dc_name:
            # If no dc_name provided, default to Kerpen
            dc_name = "Kerpen"
        
        dc_lower = dc_name.lower()
        location_map = {
            "kerpen": (NETTOMD_KERPEN_CMR_TEMPLATE_PATH, "CMR NettoMD Kerpen.xls"),
            "hodenhagen": (NETTOMD_HODENHAGEN_CMR_TEMPLATE_PATH, "CMR NettoMD Hodenhagen.xls"),
            "henstedt": (NETTOMD_HENSTEDT_ULZBURG_CMR_TEMPLATE_PATH, "CMR NettoMD Henstedt-Ulzburg.xls"),
            "hamm": (NETTOMD_HAMM_CMR_TEMPLATE_PATH, "CMR NettoMD Hamm.xls"),
            "ganderkesee": (NETTOMD_GANDERKESEE_CMR_TEMPLATE_PATH, "CMR NettoMD Ganderkesee.xls"),
            "bottrop": (NETTOMD_BOTTROP_CMR_TEMPLATE_PATH, "CMR NettoMD Bottrop.xls"),
            "krefeld": (NETTOMD_KREFELD_CMR_TEMPLATE_PATH, "CMR NettoMD Krefeld.xls"),
        }
        
        # Find matching location
        template_path = None
        output_name = None
        for key, (path, name) in location_map.items():
            if key in dc_lower:
                template_path = path
                output_name = name
                break
        
        # Default to Kerpen if not found
        if not template_path:
            template_path = NETTOMD_KERPEN_CMR_TEMPLATE_PATH
            output_name = "CMR NettoMD Kerpen.xls"
        
        output_path = WORK_DIR / f"{uuid.uuid4().hex}_{output_name}"
        try:
            # Simply copy the template file (no data filling needed for NettoMD CMR)
            import shutil
            shutil.copy2(template_path, output_path)
            
            output_bytes = output_path.read_bytes()
            
            return output_name, output_bytes
        finally:
            if output_path.exists():
                output_path.unlink()

    # Handle Rewe/Penny CMR export (with or without email)
    if export_type == "print_cmr" and ("rewe" in client_lower or "penny" in client_lower):
        output_name = "CMR Rewe-Penny.xlsm"
        output_path = WORK_DIR / f"{uuid.uuid4().hex}_{output_name}"
        try:
            if not file_name or not file_bytes:
                # No email - create empty data
                from datetime import datetime
                parsed = ParsedOrderEmail(
                    source_file="",
                    subject="",
                    sender="",
                    received_at=datetime.now(),
                    delivery_date_to_dc="",
                    leaving_date_venlo="",
                    leaving_time_venlo="",
                    orders=[],
                )
                order = CustomerOrder(
                    customer_dc=selected_client or "Rewe/Penny",
                    fatrans_dc=selected_client or "Rewe/Penny",
                    reference="",
                    items=[],
                    slotboeking_id=None,
                )
            else:
                # With email - parse it
                email_path = WORK_DIR / f"{uuid.uuid4().hex}_{Path(file_name).name}"
                email_path.write_bytes(file_bytes)
                try:
                    parsed = parse_order_email(email_path)
                    if quantity_overrides:
                        _apply_standard_quantity_overrides(parsed, quantity_overrides)
                    if export_type == "merge":
                        order = merge_orders(parsed.orders)
                    else:
                        order = parsed.orders[order_index]
                    order = _apply_client_export_rules(order, selected_client)
                finally:
                    if email_path.exists():
                        email_path.unlink()
            
            fill_rewe_penny_cmr_template(REWE_PENNY_CMR_TEMPLATE_PATH, output_path, parsed, order)

            
            output_bytes = output_path.read_bytes()

            
            return output_name, output_bytes
        finally:
            if output_path.exists():
                output_path.unlink()

    # Handle Hanos CMR export (with or without email)
    if export_type == "print_cmr" and "hanos" in client_lower:
        output_name = "CMR Hanos.xlsm"
        output_path = WORK_DIR / f"{uuid.uuid4().hex}_{output_name}"
        try:
            if not file_name or not file_bytes:
                # No email - create empty data
                from datetime import datetime
                parsed = ParsedOrderEmail(
                    source_file="",
                    subject="",
                    sender="",
                    received_at=datetime.now(),
                    delivery_date_to_dc="",
                    leaving_date_venlo="",
                    leaving_time_venlo="",
                    orders=[],
                )
                order = CustomerOrder(
                    customer_dc="Hanos",
                    fatrans_dc="Hanos",
                    reference="",
                    items=[],
                    slotboeking_id=None,
                )
            else:
                # With email - parse it
                email_path = WORK_DIR / f"{uuid.uuid4().hex}_{Path(file_name).name}"
                email_path.write_bytes(file_bytes)
                try:
                    parsed = parse_order_email(email_path)
                    if quantity_overrides:
                        _apply_standard_quantity_overrides(parsed, quantity_overrides)
                    if export_type == "merge":
                        order = merge_orders(parsed.orders)
                    else:
                        order = parsed.orders[order_index]
                    order = _apply_client_export_rules(order, selected_client)
                finally:
                    if email_path.exists():
                        email_path.unlink()
            
            fill_hanos_cmr_template(HANOS_CMR_TEMPLATE_PATH, output_path, parsed, order)

            
            output_bytes = output_path.read_bytes()

            
            return output_name, output_bytes
        finally:
            if output_path.exists():
                output_path.unlink()

    if export_type == "print_cmr" and ("carrefour fif" in client_lower or "carrefour kdc" in client_lower):
        variant = "kdc" if "kdc" in client_lower else "fif"
        template_path = KDC_CMR_TEMPLATE_PATH if variant == "kdc" else CMR_TEMPLATE_PATH
        output_name = "CMR_CARREFOUR_KDC.xlsm" if variant == "kdc" else "CMR_CARREFOUR_FIF.xlsm"
        output_path = WORK_DIR / f"{uuid.uuid4().hex}_{output_name}"
        try:
            if not file_name or not file_bytes:
                from datetime import datetime as dt

                parsed = ParsedOrderEmail(
                    source_file="",
                    subject="",
                    sender="",
                    received_at=dt.now(),
                    delivery_date_to_dc="",
                    leaving_date_venlo="",
                    leaving_time_venlo="",
                    orders=[],
                )
                order = CustomerOrder(
                    customer_dc="Carrefour KDC" if variant == "kdc" else "Carrefour FIF",
                    fatrans_dc="Carrefour",
                    reference=" + ".join(cmr_references or []),
                    items=[],
                    slotboeking_id=None,
                )
                fill_cmr_template(
                    template_path,
                    output_path,
                    parsed,
                    order,
                    references=cmr_references,
                    pallet_places=cmr_pallet_places,
                    variant=variant,
                    pakbon_items=pakbon_items,
                )
                return output_name, output_path.read_bytes()
        finally:
            if output_path.exists():
                output_path.unlink()

    if export_type == "print_cmr" and "colruyt" in client_lower and (not file_name or not file_bytes):
        output_name = "CMR_COLRUYT.xlsx"
        output_path = WORK_DIR / f"{uuid.uuid4().hex}_{output_name}"
        try:
            from datetime import datetime as dt

            parsed = ParsedOrderEmail(
                source_file="",
                subject="",
                sender="",
                received_at=dt.now(),
                delivery_date_to_dc="",
                leaving_date_venlo="",
                leaving_time_venlo="",
                orders=[],
            )
            order = CustomerOrder(
                customer_dc="Colruyt",
                fatrans_dc="Colruyt",
                reference=" + ".join(cmr_references or []),
                items=[],
                slotboeking_id=None,
            )
            fill_colruyt_cmr_template(
                COLRUYT_CMR_TEMPLATE_PATH,
                output_path,
                parsed,
                order,
                references=cmr_references,
                pallet_places=cmr_pallet_places,
                goederen_total=goederen_total,
            )
            return output_name, output_path.read_bytes()
        finally:
            if output_path.exists():
                output_path.unlink()

    if not file_name or not file_bytes:
        raise ValueError("Email file is required for this export type")
    
    email_path = WORK_DIR / f"{uuid.uuid4().hex}_{Path(file_name).name}"
    email_path.write_bytes(file_bytes)
    try:
        # Handle explicit netto_md export type
        if export_type == "netto_md":
            try:
                netto_md = parse_netto_md_email(email_path)
                if netto_md is None:
                    # Try to give more helpful error message
                    raise ValueError(
                        "Could not parse email as NettoMD format. "
                        "Please ensure the email contains NettoMD order data with the expected structure "
                        "(headers: Delivery name, Sales order, Item number, Item description, Quantity, Unit)."
                    )
                if quantity_overrides:
                    _apply_netto_md_quantity_overrides(netto_md, quantity_overrides)
                output_path = WORK_DIR / f"{uuid.uuid4().hex}_NettoMD_Orderpicking.xlsx"
                template_path = resolve_netto_md_template_path()
                if not template_path.exists():
                    raise ValueError(f"NettoMD template not found at: {template_path}")
                write_netto_md_excel(template_path, output_path, netto_md)
                from datetime import datetime as dt
                export_date = dt.now().strftime("%d-%m-%Y")
                output_bytes = output_path.read_bytes()
                output_name = f"NettoMD_Orderpicking {export_date}.xlsx"
                return output_name, output_bytes
            except ValueError:
                # Re-raise ValueError with original message
                raise
            except Exception as e:
                raise ValueError(f"NettoMD export failed: {str(e)}")
        
        # Handle explicit special export type
        if export_type == "special":
            special = parse_havi_uien_email(email_path)
            if special is None:
                raise ValueError("Could not parse email as Havi Uien format. Please check the email format.")
            if quantity_overrides:
                _apply_special_quantity_overrides(special, quantity_overrides)
            _apply_havi_uien_manual_details(special, havi_uien_article, havi_uien_description)
            output_path = WORK_DIR / f"{uuid.uuid4().hex}_HAVI_DE_UIEN.xlsx"
            write_havi_uien_excel(output_path, special)
            output_bytes = output_path.read_bytes()
            return "HAVI_DE_UIEN.xlsx", output_bytes
        
        # Auto-detect special formats for other export types
        special = parse_havi_uien_email(email_path)
        if special is not None:
            if quantity_overrides:
                _apply_special_quantity_overrides(special, quantity_overrides)
            _apply_havi_uien_manual_details(special, havi_uien_article, havi_uien_description)
            output_path = WORK_DIR / f"{uuid.uuid4().hex}_HAVI_DE_UIEN.xlsx"
            write_havi_uien_excel(output_path, special)
            output_bytes = output_path.read_bytes()
            return "HAVI_DE_UIEN.xlsx", output_bytes

        # Check NettoMD ONLY if very specific indicators are present
        # This prevents NettoMD from interfering with other clients
        if _is_definitely_netto_md_email(email_path):
            netto_md = parse_netto_md_email(email_path)
            if netto_md is not None:
                if quantity_overrides:
                    _apply_netto_md_quantity_overrides(netto_md, quantity_overrides)
                output_path = WORK_DIR / f"{uuid.uuid4().hex}_NettoMD_Orderpicking.xlsx"
                write_netto_md_excel(resolve_netto_md_template_path(), output_path, netto_md)
                export_date = datetime.now().strftime("%d-%m-%Y")
                output_bytes = output_path.read_bytes()
                return f"NettoMD_Orderpicking {export_date}.xlsx", output_bytes

        parsed = repair_parsed_texts(parse_order_email(email_path))
        if quantity_overrides:
            _apply_standard_quantity_overrides(parsed, quantity_overrides)
        if export_type == "merge":
            order = merge_orders(select_orders_by_indexes(parsed.orders, merge_order_indexes, order_index))
        else:
            order = parsed.orders[order_index]
        order = _apply_client_export_rules(order, selected_client)

        client_lower = (selected_client or "").strip().lower()
        if export_type == "print_cmr":
            base_name = build_export_name(order, selected_client).removesuffix(".xlsx")
            if "colruyt" in client_lower:
                output_name = f"CMR_{base_name}.xlsx"
            elif "denemark" in client_lower or "denmark" in client_lower:
                output_name = "CMR Denemark.xls"
            elif "edeka laatzen" in client_lower:
                output_name = f"CMR_{base_name}.xlsx"
            else:
                output_name = f"CMR_{base_name}.xlsm"
        else:
            output_name = build_export_name(order, selected_client)
        output_path = WORK_DIR / f"{uuid.uuid4().hex}_{output_name}"
        if export_type == "print_cmr":
            if "colruyt" in client_lower:
                fill_colruyt_cmr_template(
                    COLRUYT_CMR_TEMPLATE_PATH,
                    output_path,
                    parsed,
                    order,
                    references=cmr_references,
                    pallet_places=cmr_pallet_places,
                    goederen_total=goederen_total,
                )
            elif "denemark" in client_lower or "denmark" in client_lower:
                fill_denemark_cmr_template(
                    DENEMARK_CMR_TEMPLATE_PATH,
                    output_path,
                    parsed,
                    order,
                )
            elif "edeka laatzen" in client_lower:
                fill_edeka_laatzen_cmr_template(
                    EDEKA_LAATZEN_CMR_TEMPLATE_PATH,
                    output_path,
                    parsed,
                    order,
                )
            elif "havi de" in client_lower or "havi de saturday" in client_lower:
                fill_havi_de_cmr_template(
                    HAVI_DE_CMR_TEMPLATE_PATH,
                    output_path,
                    parsed,
                    order,
                )
            elif "havi be" in client_lower:
                fill_havi_be_cmr_template(
                    HAVI_BE_CMR_TEMPLATE_PATH,
                    output_path,
                    parsed,
                    order,
                    references=cmr_references,
                )
            elif "havi nl" in client_lower:
                fill_havi_nl_cmr_template(
                    HAVI_NL_CMR_TEMPLATE_PATH,
                    output_path,
                    parsed,
                    order,
                    references=cmr_references,
                )
            else:
                chosen_cmr_template = CMR_TEMPLATE_PATH
                variant = "fif"
                if "kdc" in client_lower:
                    chosen_cmr_template = KDC_CMR_TEMPLATE_PATH
                    variant = "kdc"
                
                fill_cmr_template(
                    chosen_cmr_template,
                    output_path,
                    parsed,
                    order,
                    references=cmr_references,
                    pallet_places=cmr_pallet_places,
                    variant=variant,
                    pakbon_items=pakbon_items,
                )
            
            # Return the CMR file
            output_bytes = output_path.read_bytes()
            return output_name, output_bytes
        else:
            template_path = _resolve_standard_template(parsed, order, export_sheet, selected_client)
            fill_template(template_path, output_path, parsed, order, worksheet_name=export_sheet)

            output_bytes = output_path.read_bytes()

            return output_name, output_bytes
    finally:
        if email_path.exists():
            email_path.unlink()
        if "output_path" in locals() and output_path.exists():
            output_path.unlink()


def export_leverschema_results(
    leverschema_results: dict[str, Any],
    selected_sheet: str,
    session_date: str | None = None,
) -> tuple[str, bytes]:
    WORK_DIR.mkdir(parents=True, exist_ok=True)
    extension = ".xlsm" if LEVERSCHEMA_TEMPLATE_PATH.exists() else ".xlsx"
    output_path = WORK_DIR / f"{uuid.uuid4().hex}_Leverschema_OOH{extension}"
    try:
        export_leverschema_workbook(
            LEVERSCHEMA_TEMPLATE_PATH,
            output_path,
            leverschema_results,
            selected_sheet,
            session_date=session_date,
        )
        safe_sheet = re.sub(r"[^A-Za-z0-9_-]+", "_", selected_sheet).strip("_") or "Leverschema"
        return f"{safe_sheet}{extension}", output_path.read_bytes()
    finally:
        if output_path.exists():
            output_path.unlink()


def export_preview_data(
    preview: dict[str, Any],
    export_type: str,
    order_index: int = 0,
    quantity_overrides: dict[str, Any] | None = None,
    export_sheet: str | None = None,
    selected_client: str | None = None,
    merge_order_indexes: list[int] | None = None,
) -> tuple[str, bytes]:
    if export_type not in {"selected", "merge"}:
        raise ValueError("Preview export is only available for selected and merge exports.")

    parsed = repair_parsed_texts(parsed_email_from_preview(preview))
    if quantity_overrides:
        _apply_standard_quantity_overrides(parsed, quantity_overrides)

    if export_type == "merge":
        order = merge_orders(select_orders_by_indexes(parsed.orders, merge_order_indexes, order_index))
    else:
        order = parsed.orders[order_index]
    order = _apply_client_export_rules(order, selected_client)

    output_name = build_export_name(order, selected_client)
    output_path = WORK_DIR / f"{uuid.uuid4().hex}_{output_name}"
    try:
        fill_template(
            _resolve_standard_template(parsed, order, export_sheet, selected_client),
            output_path,
            parsed,
            order,
            worksheet_name=export_sheet,
        )
        return output_name, output_path.read_bytes()
    finally:
        if output_path.exists():
            output_path.unlink()


def parsed_email_from_preview(preview: dict[str, Any]) -> ParsedOrderEmail:
    raw_orders = preview.get("orders")
    if not isinstance(raw_orders, list) or not raw_orders:
        raise ValueError("No saved orders are available to export.")

    orders: list[CustomerOrder] = []
    for raw_order in raw_orders:
        if not isinstance(raw_order, dict):
            continue
        raw_items = raw_order.get("items")
        if not isinstance(raw_items, list):
            raw_items = []
        items = [
            OrderItem(
                article_number=repair_text_value(str(item.get("primary") or "").strip()),
                description=repair_text_value(str(item.get("secondary") or "").strip()),
                quantity_boxes=_coerce_quantity(item.get("quantity")) or 0,
                unit=repair_text_value(str(item.get("unit") or "Collo").strip() or "Collo"),
            )
            for item in raw_items
            if isinstance(item, dict)
        ]
        orders.append(
            CustomerOrder(
                customer_dc=repair_text_value(str(raw_order.get("customer") or "").strip()),
                fatrans_dc=repair_text_value(str(raw_order.get("fatrans") or "").strip()),
                reference=repair_text_value(str(raw_order.get("reference") or "").strip()),
                items=items,
                slotboeking_id=None,
            )
        )

    if not orders:
        raise ValueError("No saved orders are available to export.")

    return ParsedOrderEmail(
        source_file=repair_text_value("OrderFlow saved session"),
        subject=repair_text_value("OrderFlow saved session"),
        sender="",
        received_at=datetime.now(),
        delivery_date_to_dc=repair_text_value(str(preview.get("deliveryDate") or "").strip()),
        leaving_date_venlo="",
        leaving_time_venlo="",
        orders=orders,
    )


def export_laadschema_data(
    laadschema_data: dict[str, Any],
    selected_day: str,
    selected_date: str | None = None,
    selected_week: str | int | None = None,
    custom_trucks: dict[str, Any] | None = None,
) -> tuple[str, bytes]:
    """Export Laadschema data to Excel file."""
    from openpyxl import load_workbook
    from datetime import datetime
    
    if custom_trucks is None:
        custom_trucks = {}
    
    WORK_DIR.mkdir(parents=True, exist_ok=True)
    extension = ".xlsx"
    output_path = WORK_DIR / f"{uuid.uuid4().hex}_Laadschema{extension}"
    
    try:
        # Load the template
        if not LAADSCHEMA_TEMPLATE_PATH.exists():
            raise FileNotFoundError(f"Laadschema template not found: {LAADSCHEMA_TEMPLATE_PATH}")
        
        wb = load_workbook(LAADSCHEMA_TEMPLATE_PATH)
        
        # Process ALL days that have data
        days_to_process = ['Maandag', 'Dinsdag', 'Woensdag', 'Donderdag', 'Vrijdag', 'Zaterdag', 'Zondag']
        
        for day_name in days_to_process:
            # Skip if no data for this day
            if day_name not in laadschema_data or not laadschema_data[day_name]:
                continue
            
            # Select the sheet for this day
            if day_name not in wb.sheetnames:
                continue
            
            ws = wb[day_name]
            
            # Update date and week only for the selected day
            if day_name == selected_day:
                if selected_date:
                    try:
                        date_obj = datetime.strptime(selected_date, "%Y-%m-%d")
                        ws['A2'] = date_obj
                    except:
                        pass
                
                if selected_week:
                    try:
                        ws['B2'] = int(selected_week)
                    except:
                        pass
            
            # Get data for this day
            day_data = laadschema_data.get(day_name, {})
            
            # Original client order in Excel (and web app before sorting)
            # Skip column D (Rewe/Penny 16:00) - start from column E
            clients_original = [
                'Hanos', 'McD Duisburg', 'MCD BE', 'Carrefour FIF',
                'McD NL', 'McD DE', 'McD UK', 'DK auto 1', 'DK auto 2',
                'Rewe / Penny', 'Rewe / Penny', 'Rewe / Penny', 'Rewe / Penny',
                'Netto MD', 'Carrefour KDC', 'Colruyt', 'Hello Fresh'
            ]
            
            planned_times_original = [
                '09:00', '10:00', '11:00', '12:30', '13:00', '15:30', '16:00',
                '16:30', '17:30', '17:00', '17:30', '18:00', '19:00', '22:00',
                '04:00', '01:30', '03:00'
            ]
            
            # Add custom trucks for this day
            if day_name in custom_trucks and isinstance(custom_trucks[day_name], list):
                for truck in custom_trucks[day_name]:
                    if isinstance(truck, dict) and 'clientName' in truck and 'plannedTime' in truck:
                        clients_original.append(truck['clientName'])
                        planned_times_original.append(truck['plannedTime'])
            
            # Create sorted order (same logic as in renderLaadschemaTable)
            client_indices = []
            for idx, client in enumerate(clients_original):
                client_indices.append({
                    'idx': idx,
                    'client': client,
                    'time': planned_times_original[idx]
                })
            
            after_midnight_clients = ['Carrefour KDC', 'Colruyt', 'Hello Fresh']
            before_midnight = [c for c in client_indices if c['client'] not in after_midnight_clients]
            after_midnight = [c for c in client_indices if c['client'] in after_midnight_clients]
            
            # Sort by time with stable sort (maintain original order for same times)
            def time_to_minutes(time_str):
                parts = time_str.split(':')
                return int(parts[0]) * 60 + int(parts[1])
            
            before_midnight.sort(key=lambda c: (time_to_minutes(c['time']), c['idx']))
            after_midnight.sort(key=lambda c: (time_to_minutes(c['time']), c['idx']))
            
            sorted_indices = before_midnight + after_midnight
            
            # Create mapping: sorted_col_index -> original_col_index
            col_mapping = {}
            for sorted_idx, item in enumerate(sorted_indices):
                col_mapping[sorted_idx] = item['idx']
            
            # Map row indices from web app to Excel rows
            row_mapping = {
                0: 4,   # Opmerking
                1: 7,   # Bon afgemeld
                2: 8,   # Fysieke controle
                3: 9,   # Gemeten temperatuur
                4: 10,  # Ingestelde temperatuur
                5: 11,  # Logger geactiveerd
                6: 12,  # Wielkeggen geplaatst
                7: 13,  # Tijdstip gepland vertrek
                8: 14,  # Tijdstip werkelijk vertrek
                9: 15,  # Tijdstip aankomst chauffeur
                10: 17, # Reden vertraging
                11: 18, # Overige producten
                12: 19, # Totaal aantal plaatsen
            }
            
            # Add custom truck columns to Excel (after the last standard column)
            num_standard_clients = 17  # Number of standard clients
            if day_name in custom_trucks and isinstance(custom_trucks[day_name], list):
                from copy import copy
                
                # Use the last standard column as a reference for formatting (Hello Fresh - column V = 22)
                reference_col = 21  # Column V (0-indexed: 21)
                
                for truck_idx, truck in enumerate(custom_trucks[day_name]):
                    if isinstance(truck, dict) and 'clientName' in truck and 'plannedTime' in truck:
                        # Find the position of this custom truck in the sorted list
                        custom_truck_original_idx = num_standard_clients + truck_idx
                        
                        # Find where this truck appears in the sorted list
                        sorted_position = None
                        for sorted_idx, item in enumerate(sorted_indices):
                            if item['idx'] == custom_truck_original_idx:
                                sorted_position = sorted_idx
                                break
                        
                        if sorted_position is not None:
                            # Excel column for this custom truck
                            excel_col = custom_truck_original_idx + 5
                            
                            # Copy formatting from reference column to all rows
                            for row_num in range(3, 20):  # Rows 3-19 (header to last data row)
                                source_cell = ws.cell(row=row_num, column=reference_col)
                                target_cell = ws.cell(row=row_num, column=excel_col)
                                
                                # Copy cell formatting
                                if source_cell.has_style:
                                    target_cell.font = copy(source_cell.font)
                                    target_cell.border = copy(source_cell.border)
                                    target_cell.fill = copy(source_cell.fill)
                                    target_cell.number_format = copy(source_cell.number_format)
                                    target_cell.protection = copy(source_cell.protection)
                                    target_cell.alignment = copy(source_cell.alignment)
                            
                            # Set column width same as reference column
                            ref_col_letter = ws.cell(row=1, column=reference_col).column_letter
                            target_col_letter = ws.cell(row=1, column=excel_col).column_letter
                            if ref_col_letter in ws.column_dimensions:
                                ws.column_dimensions[target_col_letter].width = ws.column_dimensions[ref_col_letter].width
                            
                            # Add column header (row 3 in Excel)
                            ws.cell(row=3, column=excel_col).value = truck['clientName']
                            
                            # Add planned time (row 13 in Excel)
                            try:
                                time_parts = truck['plannedTime'].split(':')
                                if len(time_parts) == 2:
                                    from datetime import time
                                    ws.cell(row=13, column=excel_col).value = time(int(time_parts[0]), int(time_parts[1]))
                            except:
                                ws.cell(row=13, column=excel_col).value = truck['plannedTime']
                            
                            # Add default values for specific rows (same as standard columns)
                            ws.cell(row=10, column=excel_col).value = 3  # Ingestelde temperatuur
                            ws.cell(row=11, column=excel_col).value = 'x'  # Logger geactiveerd
                            ws.cell(row=19, column=excel_col).value = 0  # Totaal aantal plaatsen
            
            # Fill in the data
            for row_idx_str, row_data in day_data.items():
                row_idx = int(row_idx_str)
                if row_idx not in row_mapping:
                    continue
                
                excel_row = row_mapping[row_idx]
                
                for col_idx_str, value in row_data.items():
                    sorted_col_idx = int(col_idx_str)
                    
                    # Map sorted column index back to original column index
                    if sorted_col_idx not in col_mapping:
                        continue
                    
                    original_col_idx = col_mapping[sorted_col_idx]
                    
                    # Column E is index 5 in Excel (1-based: A=1, B=2, C=3, D=4, E=5)
                    # We start from E because we skip column D
                    excel_col = original_col_idx + 5
                    
                    cell = ws.cell(row=excel_row, column=excel_col)
                    
                    # Handle time values
                    if value and isinstance(value, str) and ':' in value and row_idx in [8, 9]:
                        try:
                            time_parts = value.split(':')
                            if len(time_parts) == 2:
                                from datetime import time
                                cell.value = time(int(time_parts[0]), int(time_parts[1]))
                            else:
                                cell.value = value
                        except:
                            cell.value = value
                    else:
                        # Convert numeric strings to numbers
                        if value and isinstance(value, str):
                            try:
                                if value.replace('.', '', 1).replace(',', '', 1).isdigit():
                                    cell.value = float(value.replace(',', '.'))
                                else:
                                    cell.value = value
                            except:
                                cell.value = value
                        else:
                            cell.value = value
        
        # Save the workbook
        wb.save(output_path)
        
        # Create filename with day and date
        safe_day = re.sub(r"[^A-Za-z0-9_-]+", "_", selected_day).strip("_") or "Laadschema"
        if selected_date:
            filename = f"Laadschema_{safe_day}_{selected_date}{extension}"
        else:
            filename = f"Laadschema_{safe_day}{extension}"
        
        return filename, output_path.read_bytes()
    finally:
        if output_path.exists():
            output_path.unlink()


def read_uploaded_file(headers: dict[str, str], body: bytes) -> tuple[str | None, bytes | None]:
    file_name, file_bytes, _ = read_form_payload(headers, body)
    return file_name, file_bytes


def read_form_payload(headers: dict[str, str], body: bytes) -> tuple[str | None, bytes | None, dict[str, str]]:
    headers = {key.lower(): value for key, value in headers.items()}
    content_type = headers.get("content-type", "")
    if "multipart/form-data" not in content_type:
        return None, None, {}

    raw = f"Content-Type: {content_type}\r\nMIME-Version: 1.0\r\n\r\n".encode("utf-8") + body
    message = BytesParser(policy=policy.default).parsebytes(raw)
    fields: dict[str, str] = {}
    found_file_name: str | None = None
    found_file_bytes: bytes | None = None
    for part in message.iter_parts():
        if part.get_content_disposition() != "form-data":
            continue
        field_name = part.get_param("name", header="content-disposition") or ""
        filename = part.get_filename()
        if filename:
            payload = part.get_payload(decode=True)
            if payload is None:
                content = part.get_content()
                if hasattr(content, "as_bytes"):
                    payload = content.as_bytes()
                elif isinstance(content, str):
                    payload = content.encode("utf-8")
                elif isinstance(content, bytes):
                    payload = content
                else:
                    raw_payload = part.get_payload(decode=False)
                    if isinstance(raw_payload, str):
                        payload = raw_payload.encode("utf-8")
                    elif isinstance(raw_payload, bytes):
                        payload = raw_payload
                    else:
                        payload = part.as_bytes()
            found_file_name = filename
            found_file_bytes = payload
        else:
            fields[field_name] = part.get_content()
    return found_file_name, found_file_bytes, fields


def json_response(handler, payload: dict[str, Any], status: int = 200) -> None:
    data = json.dumps(payload).encode("utf-8")
    handler.send_response(status)
    handler.send_header("Content-Type", "application/json; charset=utf-8")
    handler.send_header("Content-Length", str(len(data)))
    handler.end_headers()
    handler.wfile.write(data)


def serialize_standard(parsed: ParsedOrderEmail) -> dict[str, Any]:
    return {
        "deliveryDate": parsed.delivery_date_to_dc,
        "customerCount": len(parsed.orders),
        "orders": [
            {
                "label": build_order_label(order, parsed.orders),
                "customer": order.customer_dc,
                "reference": order.reference,
                "fatrans": order.fatrans_dc,
                "deliveryPoint": short_customer_name(order.customer_dc) if _is_denmark_dc(order.fatrans_dc) else "",
                "items": [
                    {
                        "primary": item.article_number,
                        "secondary": item.description,
                        "quantity": format_number(item.quantity_boxes),
                        "unit": item.unit,
                    }
                    for item in order.items
                ],
            }
            for order in parsed.orders
        ],
        "canMerge": can_merge_orders(parsed.orders),
    }


def serialize_special(data: HaviUienOrder) -> dict[str, Any]:
    return {
        "deliveryDate": data.delivery_date,
        "customerCount": 1,
        "customer": "Havi Logistics GmbH",
        "reference": "Multiple VCSO references",
        "items": [
            {
                "primary": row.destination,
                "secondary": row.reference,
                "slicesQuantity": row.slices,
                "quantity": row.cases,
                "unit": data.case_label,
            }
            for row in data.destinations
        ],
    }


def serialize_netto_md(data: NettoMdEmail) -> dict[str, Any]:
    return {
        "deliveryDate": data.delivery_date,
        "customerCount": len(data.orders),
        "dcName": data.dc_name,
        "orders": [
            {
                "label": order.customer_name,
                "customer": order.customer_name,
                "reference": order.reference or order.sales_order,
                "fatrans": data.dc_name,
                "salesOrder": order.sales_order,
                "items": [
                    {
                        "primary": item.article_number,
                        "secondary": item.description,
                        "quantity": format_number(item.quantity),
                        "unit": item.unit,
                    }
                    for item in order.items
                ],
            }
            for order in data.orders
        ],
        "canMerge": False,
    }


def format_number(value: float) -> str:
    return str(int(value)) if value.is_integer() else f"{value:.2f}"


def build_order_label(order: CustomerOrder, orders: list[CustomerOrder]) -> str:
    display_name = short_customer_name(order.customer_dc) if _is_denmark_dc(order.fatrans_dc) else order.customer_dc
    names = [item.customer_dc for item in orders]
    if names.count(order.customer_dc) > 1:
        return f"{display_name} ({order.reference})"
    return display_name


def can_merge_orders(orders: list[CustomerOrder]) -> bool:
    if len(orders) < 2:
        return False
    return len({order.customer_dc.strip().lower() for order in orders}) == 1


def merge_orders(orders: list[CustomerOrder]) -> CustomerOrder:
    first = orders[0]
    merged_items = []
    references: list[str] = []
    slotboeking_ids: list[str] = []
    for order in orders:
        merged_items.extend(order.items)
        if order.reference and order.reference not in references:
            references.append(order.reference)
        if order.slotboeking_id and order.slotboeking_id not in slotboeking_ids:
            slotboeking_ids.append(order.slotboeking_id)
    return CustomerOrder(
        customer_dc=first.customer_dc,
        fatrans_dc=first.fatrans_dc,
        reference=" + ".join(references),
        items=merged_items,
        slotboeking_id=slotboeking_ids[0] if slotboeking_ids else None,
    )


def select_orders_by_indexes(
    orders: list[CustomerOrder],
    indexes: list[int] | None,
    fallback_index: int = 0,
) -> list[CustomerOrder]:
    selected: list[CustomerOrder] = []
    seen: set[int] = set()
    for raw_index in indexes or []:
        try:
            index = int(raw_index)
        except (TypeError, ValueError):
            continue
        if index in seen or index < 0 or index >= len(orders):
            continue
        selected.append(orders[index])
        seen.add(index)
    if selected:
        return selected
    if 0 <= fallback_index < len(orders):
        return [orders[fallback_index]]
    return orders[:1]


TEXT_REPLACEMENTS = {
    "K��ln": "Köln",
    "KÃ¶ln": "Köln",
    "Gro��beeren": "Großbeeren",
    "GroÃŸbeeren": "Großbeeren",
    "K��ge": "Køge",
    "KÃ¸ge": "Køge",
    "G��nzburg": "Günzburg",
    "GÃ¼nzburg": "Günzburg",
    "H??nchen": "Hähnchen",
    "H??hnchen": "Hähnchen",
    "H��nchen": "Hähnchen",
    "H��hnchen": "Hähnchen",
    "Hï¿½ï¿½nchen": "Hähnchen",
    "HÃ¤hnchen": "Hähnchen",
    "K??se": "Käse",
    "K��se": "Käse",
    "Kï¿½ï¿½se": "Käse",
    "KÃ¤se": "Käse",
    "Ziegenk��se": "Ziegenkäse",
    "ZiegenkÃ¤se": "Ziegenkäse",
    "M��ckm��hl": "Möckmühl",
    "MÃ¶ckmÃ¼hl": "Möckmühl",
}


def repair_text_value(value: Any) -> str:
    repaired = str(value or "")
    for source, target in TEXT_REPLACEMENTS.items():
        repaired = repaired.replace(source, target)
    return repaired


def repair_order_texts(order: CustomerOrder) -> CustomerOrder:
    return replace(
        order,
        customer_dc=repair_text_value(order.customer_dc),
        fatrans_dc=repair_text_value(order.fatrans_dc),
        reference=repair_text_value(order.reference),
        items=[
            OrderItem(
                article_number=repair_text_value(item.article_number),
                description=repair_text_value(item.description),
                quantity_boxes=item.quantity_boxes,
                unit=repair_text_value(item.unit),
            )
            for item in order.items
        ],
        slotboeking_id=repair_text_value(order.slotboeking_id) if order.slotboeking_id else None,
    )


def repair_parsed_texts(parsed: ParsedOrderEmail) -> ParsedOrderEmail:
    parsed.source_file = repair_text_value(parsed.source_file)
    parsed.subject = repair_text_value(parsed.subject)
    parsed.sender = repair_text_value(parsed.sender)
    parsed.delivery_date_to_dc = repair_text_value(parsed.delivery_date_to_dc)
    parsed.leaving_date_venlo = repair_text_value(parsed.leaving_date_venlo)
    parsed.leaving_time_venlo = repair_text_value(parsed.leaving_time_venlo)
    parsed.orders = [repair_order_texts(order) for order in parsed.orders]
    return parsed


def short_customer_name(value: str) -> str:
    value = repair_text_value(value)
    cleaned = re.sub(r"^(netto|rewe|edeka)\s+", "", value.strip(), flags=re.IGNORECASE)
    cleaned = cleaned.split(" en ")[0].split(" and ")[0].strip()
    return cleaned or value


def _is_denmark_dc(value: str) -> bool:
    normalized = unicodedata.normalize("NFKD", str(value or ""))
    ascii_only = normalized.encode("ascii", "ignore").decode("ascii")
    return ascii_only.strip().lower() == "denemark"


def sanitize_filename(value: str) -> str:
    replacements = {
        "\u00f8": "o",
        "\u00d8": "O",
        "\u00e5": "a",
        "\u00c5": "A",
        "\u00e6": "ae",
        "\u00c6": "Ae",
        "\u00df": "ss",
        "\u00e4": "ae",
        "\u00c4": "Ae",
        "\u00f6": "oe",
        "\u00d6": "Oe",
        "\u00fc": "ue",
        "\u00dc": "Ue",
    }
    for source, target in replacements.items():
        value = value.replace(source, target)
    normalized = unicodedata.normalize("NFKD", value)
    ascii_only = normalized.encode("ascii", "ignore").decode("ascii")
    cleaned = re.sub(r"[\\\\/:*?\"<>|]+", "", ascii_only)
    cleaned = re.sub(r"\s+", "_", cleaned.strip())
    return cleaned or "order"


def build_export_name(order: CustomerOrder, selected_client: str | None = None) -> str:
    order = repair_order_texts(order)
    short_name = short_customer_name(order.customer_dc)
    
    # For REWE and Penny, prepend the client name
    if selected_client:
        client_lower = selected_client.lower()
        if client_lower == "rewe":
            return f"{sanitize_filename('Rewe_' + short_name)}.xlsx"
        elif client_lower == "penny":
            return f"{sanitize_filename('Penny_' + short_name)}.xlsx"
    
    # Special handling for REWE emails when fatrans_dc is REWE
    if order.fatrans_dc and order.fatrans_dc.lower() == "rewe":
        return f"{sanitize_filename('Rewe_' + short_name)}.xlsx"
    
    return f"{sanitize_filename(short_name)}.xlsx"


def _resolve_standard_template(
    parsed: ParsedOrderEmail,
    order: CustomerOrder,
    export_sheet: str | None,
    selected_client: str | None = None,
) -> Path:
    order = repair_order_texts(order)
    customer_text = normalize_text_value(order.customer_dc)
    fatrans_text = normalize_text_value(order.fatrans_dc)
    subject_text = normalize_text_value(parsed.subject)
    source_text = normalize_text_value(parsed.source_file)
    client_text = normalize_text_value(selected_client)
    combined_text = " ".join(
        part for part in (customer_text, fatrans_text, subject_text, source_text, client_text) if part
    )
    if export_sheet and ("rewe" in combined_text or "penny" in combined_text):
        return REWE_PENNY_TEMPLATE_PATH
    return TEMPLATE_PATH


def _apply_client_export_rules(order: CustomerOrder, selected_client: str | None) -> CustomerOrder:
    order = repair_order_texts(order)
    client_text = normalize_text_value(selected_client)
    if client_text == "rewe":
        # For REWE, keep the original customer_dc (location) and just set fatrans_dc
        return replace(order, fatrans_dc="Rewe")
    if client_text == "penny":
        # For PENNY, keep the original customer_dc (location) and just set fatrans_dc  
        return replace(order, fatrans_dc="Penny")
    return order


def normalize_text_value(value: str) -> str:
    normalized = unicodedata.normalize("NFKD", str(value or ""))
    return normalized.encode("ascii", "ignore").decode("ascii").lower()


def _apply_standard_quantity_overrides(parsed: ParsedOrderEmail, quantity_overrides: dict[str, Any]) -> None:
    orders_overrides = quantity_overrides.get("orders")
    if not isinstance(orders_overrides, list):
        return
    for order_index, order_override in enumerate(orders_overrides):
        if order_index >= len(parsed.orders) or not isinstance(order_override, dict):
            continue
        item_quantities = order_override.get("items")
        if not isinstance(item_quantities, list):
            continue
        for item_index, raw_quantity in enumerate(item_quantities):
            if item_index >= len(parsed.orders[order_index].items):
                continue
            quantity = _coerce_quantity(raw_quantity)
            if quantity is not None:
                parsed.orders[order_index].items[item_index].quantity_boxes = quantity


def _apply_special_quantity_overrides(special: HaviUienOrder, quantity_overrides: dict[str, Any]) -> None:
    item_quantities = quantity_overrides.get("items")
    slice_quantities = quantity_overrides.get("slices")
    if isinstance(item_quantities, list):
        for index, raw_quantity in enumerate(item_quantities):
            if index >= len(special.destinations):
                continue
            quantity = _coerce_quantity(raw_quantity)
            if quantity is not None:
                special.destinations[index].cases = int(quantity)
    if isinstance(slice_quantities, list):
        for index, raw_quantity in enumerate(slice_quantities):
            if index >= len(special.destinations):
                continue
            quantity = _coerce_quantity(raw_quantity)
            if quantity is not None:
                special.destinations[index].slices = int(quantity)


def _apply_havi_uien_manual_details(
    special: HaviUienOrder,
    article: str | None,
    description: str | None,
) -> None:
    special.manual_article = str(article or "").strip()
    special.manual_description = str(description or "").strip()


def _apply_netto_md_quantity_overrides(netto_md: NettoMdEmail, quantity_overrides: dict[str, Any]) -> None:
    orders_overrides = quantity_overrides.get("orders")
    if not isinstance(orders_overrides, list):
        return
    for order_index, order_override in enumerate(orders_overrides):
        if order_index >= len(netto_md.orders) or not isinstance(order_override, dict):
            continue
        item_quantities = order_override.get("items")
        if not isinstance(item_quantities, list):
            continue
        for item_index, raw_quantity in enumerate(item_quantities):
            if item_index >= len(netto_md.orders[order_index].items):
                continue
            quantity = _coerce_quantity(raw_quantity)
            if quantity is not None:
                netto_md.orders[order_index].items[item_index].quantity = quantity


def _coerce_quantity(value: Any) -> float | None:
    if isinstance(value, (int, float)):
        return float(value)
    if isinstance(value, str):
        cleaned = value.replace(",", ".").strip()
        if not cleaned:
            return None
        try:
            return float(cleaned)
        except ValueError:
            return None
    return None
