from __future__ import annotations

import json
import os
import tempfile
import uuid
from datetime import date, datetime
from http import HTTPStatus
from http.server import BaseHTTPRequestHandler, ThreadingHTTPServer
from io import BytesIO
from pathlib import Path
from urllib.parse import parse_qs, urlparse

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.utils import get_column_letter

from api._auth import (
    authenticate_user,
    clear_session_cookie,
    current_user_from_headers,
    set_session_cookie,
    unauthorized_response,
)
from api._shared import (
    export_laadschema_data,
    export_leverschema_results,
    export_preview_data,
    export_uploaded_email,
    json_response,
    parse_uploaded_email,
    read_form_payload,
    read_uploaded_file,
)
from api._team_state_store import load_team_state, save_team_state
from greenops_shortage_bridge import build_shortage_previews, save_previews_as_sessions
from shortage_app import (
    as_number,
    build_day_analytics_workbook,
    build_mancos_export_rows,
    build_mancos_export_workbook,
    load_day_sessions,
    load_sessions,
    mancos_export_file_name,
    mancos_total_row,
    multipart_file,
    parse_excel,
    safe_file_part,
    save_day_sessions,
    save_sessions,
    style_mancos_sheet,
)


HOST = os.environ.get("HOST", "127.0.0.1")
PORT = int(os.environ.get("PORT", "8030"))
BASE_DIR = Path(__file__).resolve().parent
PUBLIC_DIR = BASE_DIR / "public"
SHORTAGES_DIR = PUBLIC_DIR / "shortages"


STOCK_HEADER_KEYS = {
    "itemnumber": "itemNumber",
    "productname": "productName",
    "physicalinventoryhu": "quantity",
    "expirationdate": "tht",
}


def normalize_stock_header(value: object) -> str:
    return "".join(ch for ch in str(value or "").lower() if ch.isalnum())


def format_stock_cell(value: object) -> str | int | float:
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.strftime("%d/%m/%Y")
    if isinstance(value, date):
        return value.strftime("%d/%m/%Y")
    if isinstance(value, float) and value.is_integer():
        return int(value)
    return value


def parse_stock_workbook(file_name: str, raw: bytes) -> list[dict[str, object]]:
    if not file_name.lower().endswith((".xlsx", ".xlsm")):
        raise ValueError("Upload an Excel workbook (.xlsx or .xlsm).")

    workbook = load_workbook(BytesIO(raw), read_only=True, data_only=True)
    try:
        sheet = workbook[workbook.sheetnames[0]]
        header_columns: dict[str, int] = {}
        header_row = 0
        for row_index, row in enumerate(sheet.iter_rows(min_row=1, max_row=min(sheet.max_row or 1, 30), values_only=True), start=1):
            found: dict[str, int] = {}
            for column_index, value in enumerate(row):
                mapped = STOCK_HEADER_KEYS.get(normalize_stock_header(value))
                if mapped:
                    found[mapped] = column_index
            if set(found) == {"itemNumber", "productName", "quantity", "tht"}:
                header_columns = found
                header_row = row_index
                break

        if not header_columns:
            raise ValueError("Could not find the Stock columns in this workbook.")

        items: list[dict[str, object]] = []
        for row in sheet.iter_rows(min_row=header_row + 1, values_only=True):
            item = {
                "itemNumber": format_stock_cell(row[header_columns["itemNumber"]] if header_columns["itemNumber"] < len(row) else ""),
                "productName": format_stock_cell(row[header_columns["productName"]] if header_columns["productName"] < len(row) else ""),
                "quantity": format_stock_cell(row[header_columns["quantity"]] if header_columns["quantity"] < len(row) else ""),
                "tht": format_stock_cell(row[header_columns["tht"]] if header_columns["tht"] < len(row) else ""),
            }
            if any(str(value or "").strip() for value in item.values()):
                items.append(item)
        return items
    finally:
        workbook.close()


def build_stock_export(items: list[dict[str, object]]) -> bytes:
    if not items:
        raise ValueError("There is no Stock data to export.")

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Stock"
    headers = ["Item number", "Product name", "Quantity", "THT"]

    dark_green = "0B6B43"
    soft_green = "E7F4EC"
    header_fill = PatternFill("solid", fgColor=dark_green)
    info_fill = PatternFill("solid", fgColor=soft_green)
    border_side = Side(style="thin", color="9CCCB0")
    cell_border = Border(left=border_side, right=border_side, top=border_side, bottom=border_side)

    sheet.append(["Stock"])
    sheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
    title_cell = sheet.cell(row=1, column=1)
    title_cell.font = Font(size=22, bold=True, color="FFFFFF")
    title_cell.fill = header_fill
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    sheet.row_dimensions[1].height = 34

    sheet.append(["Client", "Stock", "Export date", datetime.now().strftime("%d/%m/%Y")])
    for cell in sheet[2]:
        cell.font = Font(size=20, bold=cell.column in {1, 3}, color="08241A")
        cell.fill = info_fill
        cell.alignment = Alignment(vertical="center")
        cell.border = cell_border
    sheet.row_dimensions[2].height = 32

    sheet.append(headers)
    header_row = 3
    for cell in sheet[header_row]:
        cell.font = Font(size=20, bold=True, color="FFFFFF")
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = cell_border
    sheet.row_dimensions[header_row].height = 34

    for item in items:
        sheet.append([
            item.get("itemNumber", ""),
            item.get("productName", ""),
            item.get("quantity", ""),
            item.get("tht", ""),
        ])
        current_row = sheet.max_row
        for cell in sheet[current_row]:
            cell.font = Font(size=20, color="08241A")
            cell.alignment = Alignment(vertical="center", wrap_text=False)
            cell.border = cell_border
        sheet.row_dimensions[current_row].height = 32

    last_row = sheet.max_row
    table = Table(displayName="StockTable", ref=f"A{header_row}:D{last_row}")
    table.tableStyleInfo = TableStyleInfo(name="TableStyleMedium4", showFirstColumn=False, showLastColumn=False, showRowStripes=True, showColumnStripes=False)
    sheet.add_table(table)
    sheet.freeze_panes = "A4"
    longest_product_name = max(len(str(item.get("productName", "") or "")) for item in items)
    product_name_width = min(max(longest_product_name + 8, 58), 120)
    widths = {1: 24, 2: product_name_width, 3: 22, 4: 26}
    for index, width in widths.items():
        sheet.column_dimensions[get_column_letter(index)].width = width

    output = BytesIO()
    workbook.save(output)
    return output.getvalue()


def _find_existing_work_session(state: dict[str, object], session_date: str, session_name: str) -> dict[str, object] | None:
    target_name = session_name.strip().casefold()
    sessions = state.get("sessions")
    if not isinstance(sessions, list):
        return None
    matches = [
        session
        for session in sessions
        if isinstance(session, dict)
        and str(session.get("date") or "").strip() == session_date
        and str(session.get("name") or "").strip().casefold() == target_name
    ]
    if not matches:
        return None
    return sorted(
        matches,
        key=lambda entry: str(entry.get("createdAt") or entry.get("updatedAt") or ""),
        reverse=True,
    )[0]


class CombinedHandler(BaseHTTPRequestHandler):
    def do_GET(self) -> None:
        parsed = urlparse(self.path)
        path = parsed.path
        if path in {"/", "/index.html"}:
            self._serve_file(PUBLIC_DIR / "index.html")
            return
        if path in {"/styles.css", "/app.js", "/theme-manager.js", "/settings-icon.png", "/logout-icon.png"}:
            self._serve_file(PUBLIC_DIR / path.lstrip("/"))
            return
        if path in {"/shortages", "/shortages/"}:
            self._serve_file(SHORTAGES_DIR / "index.html")
            return
        if path.startswith("/shortages/"):
            self._serve_file(SHORTAGES_DIR / path.removeprefix("/shortages/"))
            return
        if path in {"/greenops", "/greenops/"}:
            self._serve_file(PUBLIC_DIR / "index.html")
            return
        if path.startswith("/greenops/"):
            self._serve_file(PUBLIC_DIR / path.removeprefix("/greenops/"))
            return
        if path == "/api/session":
            self._handle_session()
            return
        if path in {"/api/work_sessions", "/api/work-sessions"}:
            self._handle_get_work_sessions()
            return
        if path == "/api/day-sessions":
            json_response(self, {"daySessions": load_day_sessions()})
            return
        if path == "/api/sessions":
            json_response(self, {"sessions": load_sessions()})
            return
        self.send_error(HTTPStatus.NOT_FOUND, "Not found")

    def do_POST(self) -> None:
        parsed = urlparse(self.path)
        path = parsed.path
        if path == "/api/login":
            self._handle_login()
            return
        if path == "/api/logout":
            self._handle_logout()
            return
        if path in {"/api/work_sessions", "/api/work-sessions"}:
            self._handle_create_work_session()
            return
        if path == "/api/parse":
            self._handle_greenops_parse()
            return
        if path == "/api/export":
            self._handle_greenops_export(parsed.query)
            return
        if path == "/api/export_leverschema":
            self._handle_export_leverschema()
            return
        if path == "/api/export_laadschema":
            self._handle_export_laadschema()
            return
        if path == "/api/stock/parse":
            self._handle_stock_parse()
            return
        if path == "/api/stock/export":
            self._handle_stock_export()
            return
        if path == "/api/parse_pakbon":
            self._handle_parse_pakbon()
            return
        if path == "/api/orders/ingest":
            self._handle_order_ingest()
            return
        if path == "/api/shortages/parse":
            self._handle_shortage_parse()
            return
        if path == "/api/export-mancos":
            self._handle_export_mancos()
            return
        if path == "/api/export-day-analytics":
            self._handle_export_day_analytics()
            return
        if path == "/api/day-sessions":
            self._handle_save_day_session()
            return
        if path == "/api/sessions":
            self._handle_save_session()
            return
        self.send_error(HTTPStatus.NOT_FOUND, "Not found")

    def do_PUT(self) -> None:
        path = urlparse(self.path).path
        if path in {"/api/work_sessions", "/api/work-sessions"}:
            self._handle_put_work_sessions()
            return
        if path == "/api/sessions":
            self._handle_update_session()
            return
        self.send_error(HTTPStatus.NOT_FOUND, "Not found")

    def do_DELETE(self) -> None:
        path = urlparse(self.path).path
        if path == "/api/day-sessions":
            self._handle_delete_day_session()
            return
        self.send_error(HTTPStatus.NOT_FOUND, "Not found")

    def _handle_login(self) -> None:
        try:
            payload = json.loads(self._read_body().decode("utf-8"))
        except (UnicodeDecodeError, json.JSONDecodeError):
            json_response(self, {"error": "Invalid login payload."}, 400)
            return
        user = authenticate_user(str(payload.get("email", "")), str(payload.get("password", "")))
        if user is None:
            json_response(self, {"error": "Invalid email or password."}, 401)
            return
        data = json.dumps({"user": user}).encode("utf-8")
        self.send_response(HTTPStatus.OK)
        set_session_cookie(self, user)
        self.send_header("Content-Type", "application/json; charset=utf-8")
        self.send_header("Content-Length", str(len(data)))
        self.end_headers()
        self.wfile.write(data)

    def _handle_logout(self) -> None:
        data = json.dumps({"ok": True}).encode("utf-8")
        self.send_response(HTTPStatus.OK)
        clear_session_cookie(self)
        self.send_header("Content-Type", "application/json; charset=utf-8")
        self.send_header("Content-Length", str(len(data)))
        self.end_headers()
        self.wfile.write(data)

    def _handle_session(self) -> None:
        user = current_user_from_headers(dict(self.headers))
        if user is None:
            unauthorized_response(self)
            return
        json_response(self, {"user": user})

    def _handle_get_work_sessions(self) -> None:
        if current_user_from_headers(dict(self.headers)) is None:
            unauthorized_response(self)
            return
        try:
            state = load_team_state()
        except OSError as exc:
            json_response(self, {"error": f"Could not load team state: {exc}"}, 500)
            return
        json_response(self, state)

    def _handle_create_work_session(self) -> None:
        user = current_user_from_headers(dict(self.headers))
        if user is None:
            unauthorized_response(self)
            return
        try:
            payload = json.loads(self._read_body(max_size=200_000).decode("utf-8"))
        except (UnicodeDecodeError, json.JSONDecodeError, ValueError):
            json_response(self, {"error": "Invalid JSON body."}, 400)
            return
        if not isinstance(payload, dict):
            json_response(self, {"error": "Body must be a JSON object."}, 400)
            return
        session_date = str(payload.get("date") or "").strip()
        try:
            date.fromisoformat(session_date)
        except ValueError:
            json_response(self, {"error": "Please choose a valid session date."}, 400)
            return
        session_name = str(payload.get("name") or f"Plan for {session_date}").strip()
        try:
            existing_state = load_team_state()
        except OSError as exc:
            json_response(self, {"error": f"Could not load team state: {exc}"}, 500)
            return
        existing_session = _find_existing_work_session(existing_state, session_date, session_name)
        if existing_session:
            json_response(self, {"session": existing_session, **existing_state})
            return
        now = datetime.now().isoformat(timespec="seconds")
        session = {
            "id": str(uuid.uuid4()),
            "createdBy": str(user.get("email") or ""),
            "date": session_date,
            "name": session_name,
            "createdAt": now,
            "updatedAt": now,
            "workspaces": {},
        }
        try:
            save_team_state({"sessions": [session]})
            state = load_team_state()
        except OSError as exc:
            json_response(self, {"error": f"Could not persist work session: {exc}"}, 500)
            return
        if not any(entry.get("id") == session["id"] for entry in state.get("sessions", [])):
            state["sessions"] = [session, *state.get("sessions", [])]
        json_response(self, {"session": session, **state})

    def _handle_put_work_sessions(self) -> None:
        if current_user_from_headers(dict(self.headers)) is None:
            unauthorized_response(self)
            return
        try:
            payload = json.loads(self._read_body(max_size=6_000_000).decode("utf-8"))
        except (UnicodeDecodeError, json.JSONDecodeError, ValueError):
            json_response(self, {"error": "Invalid JSON body."}, 400)
            return
        if not isinstance(payload, dict):
            json_response(self, {"error": "Body must be a JSON object."}, 400)
            return
        try:
            save_team_state(payload)
        except OSError as exc:
            json_response(self, {"error": f"Could not persist team state: {exc}"}, 500)
            return
        json_response(self, {"ok": True})

    def _handle_greenops_parse(self) -> None:
        if current_user_from_headers(dict(self.headers)) is None:
            unauthorized_response(self)
            return
        body = self._read_body(max_size=12_000_000)
        file_name, file_bytes = read_uploaded_file(dict(self.headers), body)
        if not file_name or file_bytes is None:
            json_response(self, {"error": "No email file was uploaded."}, 400)
            return
        try:
            json_response(self, parse_uploaded_email(file_name, file_bytes))
        except Exception as exc:
            json_response(self, {"error": str(exc)}, 400)

    def _handle_greenops_export(self, query_string: str) -> None:
        if current_user_from_headers(dict(self.headers)) is None:
            unauthorized_response(self)
            return
        body = self._read_body(max_size=18_000_000)
        file_name, file_bytes, fields = read_form_payload(dict(self.headers), body)
        query = parse_qs(query_string)
        export_type = query.get("type", ["selected"])[0]
        selected_client = query.get("selectedClient", [""])[0].strip()
        client_lower = selected_client.lower()
        allow_without_email = export_type == "print_cmr" and any(
            needle in client_lower
            for needle in [
                "denemark",
                "denmark",
                "edeka laatzen",
                "edeka mochmuhl",
                "globus",
                "havi nl",
                "havi be",
                "havi de",
                "heeren",
                "nettomd",
                "rewe",
                "penny",
                "hanos",
                "carrefour fif",
                "carrefour kdc",
                "colruyt",
                "duisburg",
            ]
        )
        has_preview_export = bool(fields.get("previewExport", "").strip()) and export_type in {"selected", "merge"}
        if (not file_name or file_bytes is None) and not allow_without_email and not has_preview_export:
            json_response(self, {"error": "No email file was uploaded."}, 400)
            return

        try:
            order_index = int(query.get("orderIndex", ["0"])[0])
        except ValueError:
            order_index = 0
        quantity_overrides = self._json_field(fields, "quantityOverrides")
        merge_order_indexes = self._json_field(fields, "mergeOrderIndexes", fallback=[])
        preview_export = self._json_field(fields, "previewExport")
        cmr_references = self._json_field(fields, "cmrReferences", fallback=[])
        pakbon_items = self._json_field(fields, "pakbonItems")
        goederen_total = self._float_field(fields, "goederenTotal")
        try:
            if isinstance(preview_export, dict) and export_type in {"selected", "merge"}:
                output_name, output_bytes = export_preview_data(
                    preview_export,
                    export_type,
                    order_index,
                    quantity_overrides if isinstance(quantity_overrides, dict) else None,
                    query.get("exportSheet", [""])[0].strip() or None,
                    selected_client or None,
                    merge_order_indexes if isinstance(merge_order_indexes, list) else [],
                )
            else:
                output_name, output_bytes = export_uploaded_email(
                    file_name,
                    file_bytes,
                    export_type,
                    order_index,
                    quantity_overrides if isinstance(quantity_overrides, dict) else None,
                    query.get("exportSheet", [""])[0].strip() or None,
                    selected_client or None,
                    cmr_references if isinstance(cmr_references, list) else [],
                    fields.get("cmrPalletPlaces", "").strip() or None,
                    fields.get("dcName", "").strip() or None,
                    pakbon_items if isinstance(pakbon_items, list) else None,
                    goederen_total,
                    merge_order_indexes if isinstance(merge_order_indexes, list) else [],
                    fields.get("haviUienArticle", "").strip() or None,
                    fields.get("haviUienDescription", "").strip() or None,
                )
            self._send_file(output_name, output_bytes)
        except Exception as exc:
            json_response(self, {"error": str(exc)}, 400)

    def _handle_export_leverschema(self) -> None:
        if current_user_from_headers(dict(self.headers)) is None:
            unauthorized_response(self)
            return
        try:
            payload = json.loads(self._read_body(max_size=4_000_000).decode("utf-8") or "{}")
            output_name, output_bytes = export_leverschema_results(
                payload.get("results") if isinstance(payload.get("results"), dict) else {},
                str(payload.get("selectedSheet") or "").strip(),
                payload.get("sessionDate") if isinstance(payload.get("sessionDate"), str) else None,
            )
            self._send_file(output_name, output_bytes)
        except Exception as exc:
            json_response(self, {"error": str(exc)}, 400)

    def _handle_export_laadschema(self) -> None:
        if current_user_from_headers(dict(self.headers)) is None:
            unauthorized_response(self)
            return
        try:
            payload = json.loads(self._read_body(max_size=4_000_000).decode("utf-8") or "{}")
            output_name, output_bytes = export_laadschema_data(
                payload.get("data") if isinstance(payload.get("data"), dict) else {},
                str(payload.get("selectedDay") or "").strip(),
                payload.get("selectedDate") if isinstance(payload.get("selectedDate"), str) else None,
                payload.get("selectedWeek") if isinstance(payload.get("selectedWeek"), (str, int)) else None,
                payload.get("customTrucks") if isinstance(payload.get("customTrucks"), dict) else {},
            )
            self._send_file(output_name, output_bytes)
        except Exception as exc:
            json_response(self, {"error": str(exc)}, 400)

    def _handle_stock_parse(self) -> None:
        if current_user_from_headers(dict(self.headers)) is None:
            unauthorized_response(self)
            return
        try:
            body = self._read_body(max_size=18_000_000)
            file_name, file_bytes = read_uploaded_file(dict(self.headers), body)
            if not file_name or file_bytes is None:
                json_response(self, {"error": "No Stock workbook was uploaded."}, 400)
                return
            items = parse_stock_workbook(file_name, file_bytes)
            json_response(self, {"items": items, "count": len(items)})
        except Exception as exc:
            json_response(self, {"error": str(exc)}, 400)

    def _handle_stock_export(self) -> None:
        if current_user_from_headers(dict(self.headers)) is None:
            unauthorized_response(self)
            return
        try:
            payload = json.loads(self._read_body(max_size=10_000_000).decode("utf-8") or "{}")
            raw_items = payload.get("items")
            if not isinstance(raw_items, list):
                json_response(self, {"error": "Missing Stock rows to export."}, 400)
                return
            items = [item for item in raw_items if isinstance(item, dict)]
            output_bytes = build_stock_export(items)
            file_name = f"Stock - {datetime.now().strftime('%d-%m-%Y')}.xlsx"
            self._send_file(file_name, output_bytes)
        except Exception as exc:
            json_response(self, {"error": str(exc)}, 400)

    def _handle_parse_pakbon(self) -> None:
        if current_user_from_headers(dict(self.headers)) is None:
            unauthorized_response(self)
            return
        try:
            content_type = self.headers.get("content-type", "")
            body = self._read_body(max_size=20_000_000)
            raw = f"Content-Type: {content_type}\r\nMIME-Version: 1.0\r\n\r\n".encode("utf-8") + body
            from email import policy
            from email.parser import BytesParser
            from email_order_app.pakbon_parser import calculate_goederen_total, parse_multiple_pakbons

            message = BytesParser(policy=policy.default).parsebytes(raw)
            temp_dir = Path(tempfile.gettempdir()) / "pakbon_uploads"
            temp_dir.mkdir(parents=True, exist_ok=True)
            pdf_files: list[Path] = []
            for part in message.iter_parts():
                filename = part.get_filename()
                if not filename or not filename.lower().endswith(".pdf"):
                    continue
                payload = part.get_payload(decode=True)
                if payload:
                    temp_path = temp_dir / f"{uuid.uuid4().hex}_{Path(filename).name}"
                    temp_path.write_bytes(payload)
                    pdf_files.append(temp_path)
            if not pdf_files:
                json_response(self, {"error": "No PDF files uploaded"}, 400)
                return
            try:
                merged_items = parse_multiple_pakbons(pdf_files, section="Emballage")
                goederen_total = calculate_goederen_total(pdf_files)
                chep_total = sum(
                    item.quantity
                    for item in merged_items.values()
                    if item.article_number == "409" or "chep" in item.description.lower()
                )
                json_response(
                    self,
                    {
                        "items": [
                            {
                                "articleNumber": item.article_number,
                                "description": item.description,
                                "quantity": item.quantity,
                                "unit": item.unit,
                            }
                            for item in merged_items.values()
                        ],
                        "totalItems": len(merged_items),
                        "goederenTotal": goederen_total,
                        "chepTotal": chep_total,
                    },
                )
            finally:
                for temp_path in pdf_files:
                    try:
                        temp_path.unlink()
                    except OSError:
                        pass
        except Exception as exc:
            json_response(self, {"error": str(exc)}, 400)

    def _handle_order_ingest(self) -> None:
        try:
            body = self._read_body(max_size=12_000_000)
            file_name, file_bytes, fields = read_form_payload(dict(self.headers), body)
            session_date = str(fields.get("date") or "").strip()
            session_name = str(fields.get("name") or "").strip()
            work_session_id = str(fields.get("workSessionId") or "").strip()
            if not session_date:
                json_response(self, {"error": "Open a daily session before uploading an email."}, 400)
                return
            raw_preview = str(fields.get("preview") or "").strip()
            if raw_preview:
                mode = str(fields.get("mode") or "standard").strip() or "standard"
                try:
                    greenops_preview = json.loads(raw_preview)
                except json.JSONDecodeError:
                    json_response(self, {"error": "Invalid OrderFlow preview payload."}, 400)
                    return
                parsed = {"mode": mode, "preview": greenops_preview}
                file_name = file_name or "OrderFlow saved orders"
            else:
                if not file_name or file_bytes is None:
                    json_response(self, {"error": "No OrderFlow email file was uploaded."}, 400)
                    return
                parsed = parse_uploaded_email(file_name, file_bytes)
                mode = str(parsed.get("mode") or "")
                greenops_preview = parsed.get("preview")
            if not isinstance(greenops_preview, dict):
                json_response(self, {"error": "OrderFlow did not return an order preview."}, 400)
                return
            shortage_previews = build_shortage_previews(file_name, mode, greenops_preview)
            if not shortage_previews:
                json_response(self, {"error": "No order lines were found in this email."}, 400)
                return
            saved, sessions = save_previews_as_sessions(session_date, session_name, shortage_previews, work_session_id)
            json_response(
                self,
                {
                    "mode": mode,
                    "greenopsPreview": greenops_preview,
                    "savedSessions": saved,
                    "sessions": sessions,
                },
            )
        except ValueError as exc:
            json_response(self, {"error": str(exc)}, 400)
        except Exception as exc:
            json_response(self, {"error": f"Could not ingest OrderFlow email: {exc}"}, 400)

    def _handle_shortage_parse(self) -> None:
        try:
            file_name, raw = multipart_file(self.headers.get("content-type", ""), self._read_body(max_size=12_000_000))
            json_response(self, {"preview": parse_excel(file_name, raw)})
        except ValueError as exc:
            json_response(self, {"error": str(exc)}, 400)
        except Exception as exc:
            json_response(self, {"error": f"Could not read the file: {exc}"}, 400)

    def _handle_export_mancos(self) -> None:
        try:
            payload = json.loads(self._read_body(max_size=8_000_000).decode("utf-8"))
            if isinstance(payload.get("previews"), list):
                preview_entries = [entry for entry in payload.get("previews", []) if isinstance(entry, dict)]
            elif isinstance(payload.get("preview"), dict):
                preview_entries = [{"preview": payload["preview"]}]
            else:
                json_response(self, {"error": "Missing delivery point data."}, 400)
                return

            try:
                workbook = build_mancos_export_workbook(preview_entries)
            except ValueError as exc:
                json_response(self, {"error": str(exc)}, 400)
                return
            output = BytesIO()
            workbook.save(output)
            file_name = mancos_export_file_name(preview_entries)
            self._send_file(file_name, output.getvalue())
            return

            rows = build_mancos_export_rows(preview_entries)
            if not rows:
                json_response(self, {"error": "There are no Manco lines to export."}, 400)
                return
            workbook = Workbook()
            sheet = workbook.active
            sheet.title = "Mancos"
            sheet.append(["Manco´s Export"])
            if len(preview_entries) == 1:
                preview = preview_entries[0].get("preview", {}) if isinstance(preview_entries[0], dict) else {}
                sheet.append(["Client", preview.get("client", "")])
                sheet.append(["Delivery point", preview.get("deliveryPoint", "")])
            else:
                sheet.append(["Selected orders", len(preview_entries)])
                sheet.append(["Visible lines", "Manco > 0"])
            sheet.append([])
            sheet.append(["Client", "Delivery point", "Order Reference", "Article", "Description", "Ordered", "Delivered", "Manco", "Manco %"])
            for row in rows:
                sheet.append(
                    [
                        row.get("client", ""),
                        row.get("deliveryPoint", ""),
                        row.get("orderReference", ""),
                        row.get("article", ""),
                        row.get("description", ""),
                        row.get("orderedQuantity", 0),
                        row.get("deliveredQuantity", 0),
                        row.get("shortageQuantity", 0),
                        row.get("shortagePercentage", 0),
                    ]
                )
            sheet.append(mancos_total_row(rows))
            style_mancos_sheet(sheet, len(rows))
            minimum_widths = {1: 18, 2: 20, 3: 22, 4: 18, 5: 42, 6: 16, 7: 18, 8: 16, 9: 14}
            for column_index, column_cells in enumerate(sheet.iter_cols(min_col=1, max_col=9), start=1):
                max_length = max(len(str(cell.value or "")) for cell in column_cells)
                minimum_width = minimum_widths.get(column_index, 12)
                sheet.column_dimensions[get_column_letter(column_index)].width = min(max(max_length + 4, minimum_width), 46)
            output = BytesIO()
            workbook.save(output)
            if len(preview_entries) == 1:
                preview = preview_entries[0].get("preview", {}) if isinstance(preview_entries[0], dict) else {}
                client = safe_file_part(str(preview.get("client") or "client"))
                point = safe_file_part(str(preview.get("deliveryPoint") or "delivery-point"))
                file_name = f"mancos_{client}_{point}.xlsx"
            else:
                file_name = f"mancos_selected_{len(preview_entries)}_orders.xlsx"
            self._send_file(file_name, output.getvalue())
        except Exception as exc:
            json_response(self, {"error": str(exc)}, 400)

    def _handle_export_day_analytics(self) -> None:
        try:
            payload = json.loads(self._read_body(max_size=400_000).decode("utf-8") or "{}")
            session_date = str(payload.get("date") or "").strip()
            work_session_id = str(payload.get("workSessionId") or "").strip()
            havi_uien_settings = payload.get("haviUienSettings") if isinstance(payload.get("haviUienSettings"), dict) else None
            sessions = load_sessions()
            if work_session_id:
                sessions = [
                    session for session in sessions
                    if str(session.get("workSessionId") or "") == work_session_id
                ]
            workbook = build_day_analytics_workbook(sessions, session_date, havi_uien_settings)
            output = BytesIO()
            workbook.save(output)
            safe_date = session_date or datetime.now().strftime("%Y-%m-%d")
            self._send_file(f"Manco Analytics - {safe_date}.xlsx", output.getvalue())
        except Exception as exc:
            json_response(self, {"error": str(exc)}, 400)

    def _handle_save_session(self) -> None:
        try:
            payload = json.loads(self._read_body(max_size=2_000_000).decode("utf-8"))
            session_date = str(payload.get("date", "")).strip()
            parsed_date = date.fromisoformat(session_date)
            preview = payload.get("preview")
            if not isinstance(preview, dict):
                json_response(self, {"error": "Upload a file before saving the session."}, 400)
                return
            sessions = load_sessions()
            session = {
                "id": str(uuid.uuid4()),
                "date": session_date,
                "weekday": parsed_date.strftime("%A"),
                "name": str(payload.get("name") or f"{parsed_date.strftime('%A')} Manco session"),
                "workSessionId": str(payload.get("workSessionId") or "").strip(),
                "createdAt": datetime.now().isoformat(timespec="seconds"),
                "preview": preview,
            }
            sessions.insert(0, session)
            save_sessions(sessions[:1000])
            json_response(self, {"session": session, "sessions": sessions[:1000]})
        except ValueError:
            json_response(self, {"error": "Please choose a valid session date."}, 400)
        except Exception as exc:
            json_response(self, {"error": str(exc)}, 400)

    def _handle_save_day_session(self) -> None:
        try:
            payload = json.loads(self._read_body(max_size=200_000).decode("utf-8"))
            session_date = str(payload.get("date", "")).strip()
            parsed_date = date.fromisoformat(session_date)
            sessions = load_day_sessions()
            existing = next((entry for entry in sessions if entry.get("date") == session_date), None)
            if existing:
                json_response(self, {"daySession": existing, "daySessions": sessions})
                return
            day_session = {
                "id": str(uuid.uuid4()),
                "date": session_date,
                "weekday": parsed_date.strftime("%A"),
                "name": str(payload.get("name") or f"{parsed_date.strftime('%A')} Manco review"),
                "createdAt": datetime.now().isoformat(timespec="seconds"),
            }
            sessions.insert(0, day_session)
            sessions = sorted(sessions, key=lambda entry: str(entry.get("date", "")), reverse=True)[:1000]
            save_day_sessions(sessions)
            json_response(self, {"daySession": day_session, "daySessions": sessions})
        except ValueError:
            json_response(self, {"error": "Please choose a valid session date."}, 400)
        except Exception as exc:
            json_response(self, {"error": str(exc)}, 400)

    def _handle_delete_day_session(self) -> None:
        try:
            payload = json.loads(self._read_body(max_size=200_000).decode("utf-8"))
            session_id = str(payload.get("id") or "").strip()
            session_date = str(payload.get("date") or "").strip()
            if not session_id and not session_date:
                json_response(self, {"error": "Missing session id or date."}, 400)
                return
            day_sessions = load_day_sessions()
            matched_date = session_date
            for day_session in day_sessions:
                if session_id and day_session.get("id") == session_id:
                    matched_date = str(day_session.get("date") or "")
                    break
            if not matched_date:
                json_response(self, {"error": "Session date not found."}, 404)
                return
            day_sessions = [
                entry
                for entry in day_sessions
                if entry.get("id") != session_id and entry.get("date") != matched_date
            ]
            order_sessions = [entry for entry in load_sessions() if entry.get("date") != matched_date]
            save_day_sessions(day_sessions)
            save_sessions(order_sessions[:1000])
            json_response(self, {"ok": True, "daySessions": day_sessions, "sessions": order_sessions[:1000]})
        except Exception as exc:
            json_response(self, {"error": str(exc)}, 400)

    def _handle_update_session(self) -> None:
        try:
            payload = json.loads(self._read_body(max_size=2_000_000).decode("utf-8"))
            session_id = str(payload.get("id", "")).strip()
            preview = payload.get("preview")
            if not session_id or not isinstance(preview, dict):
                json_response(self, {"error": "Missing session or preview data."}, 400)
                return
            sessions = load_sessions()
            for session in sessions:
                if session.get("id") == session_id:
                    session["preview"] = preview
                    if payload.get("workSessionId"):
                        session["workSessionId"] = str(payload.get("workSessionId") or "").strip()
                    session["updatedAt"] = datetime.now().isoformat(timespec="seconds")
                    save_sessions(sessions[:1000])
                    json_response(self, {"session": session, "sessions": sessions[:1000]})
                    return
            json_response(self, {"error": "Session not found."}, 404)
        except Exception as exc:
            json_response(self, {"error": str(exc)}, 400)

    def _serve_file(self, path: Path) -> None:
        try:
            resolved = path.resolve()
            allowed_roots = [PUBLIC_DIR.resolve()]
            if not any(resolved == root or root in resolved.parents for root in allowed_roots):
                self.send_error(HTTPStatus.NOT_FOUND, "Not found")
                return
            if not resolved.is_file():
                self.send_error(HTTPStatus.NOT_FOUND, "Not found")
                return
            content_type = _content_type(resolved)
            data = resolved.read_bytes()
            self.send_response(HTTPStatus.OK)
            self.send_header("Content-Type", content_type)
            self.send_header("Content-Length", str(len(data)))
            self.end_headers()
            self.wfile.write(data)
        except OSError:
            self.send_error(HTTPStatus.NOT_FOUND, "Not found")

    def _send_file(self, file_name: str, data: bytes) -> None:
        self.send_response(HTTPStatus.OK)
        self.send_header("Content-Type", _download_content_type(file_name))
        self.send_header("Content-Disposition", f'attachment; filename="{file_name}"')
        self.send_header("Content-Length", str(len(data)))
        self.end_headers()
        self.wfile.write(data)

    def _read_body(self, max_size: int = 2_000_000) -> bytes:
        length = int(self.headers.get("content-length", "0"))
        if length <= 0 or length > max_size:
            raise ValueError("Invalid payload size.")
        return self.rfile.read(length)

    @staticmethod
    def _json_field(fields: dict[str, str], name: str, fallback: object | None = None) -> object | None:
        raw = fields.get(name, "").strip()
        if not raw:
            return fallback
        return json.loads(raw)

    @staticmethod
    def _float_field(fields: dict[str, str], name: str) -> float | None:
        raw = fields.get(name, "").strip()
        if not raw:
            return None
        return float(raw)


def _content_type(path: Path) -> str:
    suffix = path.suffix.lower()
    return {
        ".html": "text/html; charset=utf-8",
        ".css": "text/css; charset=utf-8",
        ".js": "application/javascript; charset=utf-8",
        ".png": "image/png",
        ".jpg": "image/jpeg",
        ".jpeg": "image/jpeg",
        ".svg": "image/svg+xml",
    }.get(suffix, "application/octet-stream")


def _download_content_type(file_name: str) -> str:
    lower = file_name.lower()
    if lower.endswith(".xlsm"):
        return "application/vnd.ms-excel.sheet.macroEnabled.12"
    if lower.endswith(".xlsx"):
        return "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    return "application/octet-stream"


def main() -> None:
    server = ThreadingHTTPServer((HOST, PORT), CombinedHandler)
    print(f"OrderFlow running at http://{HOST}:{PORT}")
    server.serve_forever()


if __name__ == "__main__":
    main()
