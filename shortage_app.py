from __future__ import annotations

import csv
import json
import os
import re
import urllib.error
import urllib.request
import uuid
from datetime import date, datetime
from http import HTTPStatus
from http.server import BaseHTTPRequestHandler, ThreadingHTTPServer
from io import BytesIO, StringIO
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.worksheet.table import Table, TableStyleInfo

from api._cloud_state_store import load_json_state, save_json_state, supabase_enabled

HOST = os.environ.get("HOST", "127.0.0.1")
PORT = int(os.environ.get("PORT", "8020"))
BASE_DIR = Path(__file__).parent
STATIC_DIR = BASE_DIR / "shortage_web"
STORAGE_PATH = BASE_DIR / "data" / "shortage_sessions.json"
DAY_SESSIONS_PATH = BASE_DIR / "data" / "shortage_day_sessions.json"
REDIS_SHORTAGE_SESSIONS_KEY = "greenops:shortage_sessions:v1"
REDIS_SHORTAGE_DAY_SESSIONS_KEY = "greenops:shortage_day_sessions:v1"
SUPABASE_SHORTAGE_SESSIONS_KEY = "shortage_sessions:v1"
SUPABASE_SHORTAGE_DAY_SESSIONS_KEY = "shortage_day_sessions:v1"


CLIENT_PROFILES = [
    {
        "client": "HAVI",
        "delivery_points": {
            "DC Duisburg": ["duisburg"],
            "DC Wunstorf": ["wunstorf"],
            "DC Neu Wulmstorf": ["neu wulmstorf", "neuwulmstorf"],
            "HAVI NL": ["havi nl", "netherlands", "nederland"],
            "HAVI BE": ["havi be", "belgium", "belgie"],
            "HAVI UK": ["havi uk", "united kingdom", "uk limited"],
        },
        "aliases": ["havi", "havi logistics"],
    },
    {
        "client": "Carrefour",
        "delivery_points": {
            "FIF": ["fif"],
            "KDC": ["kdc"],
        },
        "aliases": ["carrefour"],
    },
    {
        "client": "Colruyt",
        "delivery_points": {
            "Dassenveld": ["dassenveld"],
        },
        "aliases": ["colruyt"],
    },
    {
        "client": "Edeka",
        "delivery_points": {
            "Laatzen": ["laatzen"],
            "Mockmuhl": ["mockmuhl", "möckmühl", "moeckmuehl"],
        },
        "aliases": ["edeka"],
    },
    {
        "client": "Denemark",
        "delivery_points": {
            "Brabrand": ["brabrand"],
            "Køge": ["køge", "koge", "koege"],
        },
        "aliases": ["denemark", "denmark"],
    },
    {
        "client": "NettoMD",
        "delivery_points": {
            "Kerpen": ["kerpen"],
            "Hodenhagen": ["hodenhagen"],
            "Henstedt": ["henstedt", "henstedt-ulzburg"],
            "Hamm": ["hamm"],
            "Ganderkesee": ["ganderkesee"],
            "Bottrop": ["bottrop"],
            "Krefeld": ["krefeld"],
        },
        "aliases": ["nettomd", "netto md", "netto"],
    },
    {"client": "Rewe", "delivery_points": {}, "aliases": ["rewe"]},
    {"client": "Penny", "delivery_points": {}, "aliases": ["penny"]},
    {"client": "Globus", "delivery_points": {}, "aliases": ["globus"]},
    {"client": "Hanos", "delivery_points": {}, "aliases": ["hanos"]},
    {"client": "Heeren", "delivery_points": {}, "aliases": ["heeren", "herren"]},
    {"client": "HelloFresh", "delivery_points": {}, "aliases": ["hellofresh", "hello fresh"]},
]


ARTICLE_HEADERS = ["article", "articlenumber", "articleno", "item", "itemnumber", "sku", "productcode"]
DESCRIPTION_HEADERS = ["description", "product", "itemdescription", "name"]
ORDERED_HEADERS = ["ordered", "orderquantity", "orderedquantity", "qtyordered", "quantity", "orderqty", "orderedqty"]
DELIVERED_HEADERS = ["delivered", "available", "supplied", "shipped", "picked", "deliveryquantity"]
SHORTAGE_HEADERS = ["manco", "shortage", "short", "missing", "difference", "backorder", "notdelivered"]


def load_sessions() -> list[dict[str, Any]]:
    if supabase_enabled():
        data = load_json_state(SUPABASE_SHORTAGE_SESSIONS_KEY)
        if isinstance(data, list):
            return [normalize_saved_session(entry) for entry in data if isinstance(entry, dict)]
        return []

    url, token = redis_rest_credentials()
    if url and token:
        raw = upstash_get(url, token, REDIS_SHORTAGE_SESSIONS_KEY)
        if raw:
            try:
                data = json.loads(raw)
            except json.JSONDecodeError:
                data = []
            if isinstance(data, list):
                return [normalize_saved_session(entry) for entry in data if isinstance(entry, dict)]
        return []

    if not STORAGE_PATH.is_file():
        return []
    try:
        data = json.loads(STORAGE_PATH.read_text(encoding="utf-8"))
    except (OSError, json.JSONDecodeError):
        return []
    if not isinstance(data, list):
        return []
    return [normalize_saved_session(entry) for entry in data if isinstance(entry, dict)]


def save_sessions(sessions: list[dict[str, Any]]) -> None:
    if supabase_enabled():
        save_json_state(SUPABASE_SHORTAGE_SESSIONS_KEY, sessions)
        return

    url, token = redis_rest_credentials()
    if url and token:
        upstash_set(url, token, REDIS_SHORTAGE_SESSIONS_KEY, json.dumps(sessions, ensure_ascii=False, separators=(",", ":")))
        return

    STORAGE_PATH.parent.mkdir(parents=True, exist_ok=True)
    tmp = STORAGE_PATH.with_suffix(".json.tmp")
    tmp.write_text(json.dumps(sessions, ensure_ascii=False, indent=2), encoding="utf-8")
    tmp.replace(STORAGE_PATH)


def load_day_sessions() -> list[dict[str, Any]]:
    if supabase_enabled():
        data = load_json_state(SUPABASE_SHORTAGE_DAY_SESSIONS_KEY)
        if isinstance(data, list):
            return [entry for entry in data if isinstance(entry, dict)]
        return []

    url, token = redis_rest_credentials()
    if url and token:
        raw = upstash_get(url, token, REDIS_SHORTAGE_DAY_SESSIONS_KEY)
        if raw:
            try:
                data = json.loads(raw)
            except json.JSONDecodeError:
                data = []
            if isinstance(data, list):
                return [entry for entry in data if isinstance(entry, dict)]
        return []

    if not DAY_SESSIONS_PATH.is_file():
        return []
    try:
        data = json.loads(DAY_SESSIONS_PATH.read_text(encoding="utf-8"))
    except (OSError, json.JSONDecodeError):
        return []
    if not isinstance(data, list):
        return []
    return [entry for entry in data if isinstance(entry, dict)]


def save_day_sessions(sessions: list[dict[str, Any]]) -> None:
    if supabase_enabled():
        save_json_state(SUPABASE_SHORTAGE_DAY_SESSIONS_KEY, sessions)
        return

    url, token = redis_rest_credentials()
    if url and token:
        upstash_set(url, token, REDIS_SHORTAGE_DAY_SESSIONS_KEY, json.dumps(sessions, ensure_ascii=False, separators=(",", ":")))
        return

    DAY_SESSIONS_PATH.parent.mkdir(parents=True, exist_ok=True)
    tmp = DAY_SESSIONS_PATH.with_suffix(".json.tmp")
    tmp.write_text(json.dumps(sessions, ensure_ascii=False, indent=2), encoding="utf-8")
    tmp.replace(DAY_SESSIONS_PATH)


def redis_rest_credentials() -> tuple[str, str]:
    url = (
        os.environ.get("UPSTASH_REDIS_REST_URL")
        or os.environ.get("KV_REST_API_URL")
        or ""
    ).strip()
    token = (
        os.environ.get("UPSTASH_REDIS_REST_TOKEN")
        or os.environ.get("KV_REST_API_TOKEN")
        or ""
    ).strip()
    return url, token


def upstash_get(url: str, token: str, key: str) -> str | None:
    from urllib.parse import quote

    endpoint = url.rstrip("/") + "/get/" + quote(key, safe="")
    req = urllib.request.Request(endpoint, headers={"Authorization": f"Bearer {token}"})
    try:
        with urllib.request.urlopen(req, timeout=25) as resp:
            body = json.loads(resp.read().decode("utf-8"))
    except (urllib.error.URLError, urllib.error.HTTPError, json.JSONDecodeError, TimeoutError):
        return None
    if isinstance(body, dict) and body.get("error"):
        return None
    result = body.get("result")
    if result is None:
        return None
    if isinstance(result, str):
        return result
    return json.dumps(result)


def upstash_set(url: str, token: str, key: str, value: str) -> None:
    from urllib.parse import quote

    endpoint = url.rstrip("/") + "/set/" + quote(key, safe="")
    req = urllib.request.Request(
        endpoint,
        data=value.encode("utf-8"),
        headers={
            "Authorization": f"Bearer {token}",
            "Content-Type": "application/json; charset=utf-8",
        },
        method="POST",
    )
    try:
        with urllib.request.urlopen(req, timeout=45) as resp:
            body = json.loads(resp.read().decode("utf-8"))
    except urllib.error.HTTPError as exc:
        detail = ""
        try:
            detail = exc.read().decode("utf-8", errors="replace")[:300]
        except OSError:
            pass
        raise OSError(f"Upstash HTTP {exc.code} {detail}") from exc
    except urllib.error.URLError as exc:
        raise OSError(str(exc)) from exc
    if not isinstance(body, dict):
        raise OSError("Unexpected Upstash response")
    if body.get("error"):
        raise OSError(str(body["error"]))


def normalize_saved_session(session: dict[str, Any]) -> dict[str, Any]:
    session = repair_display_texts(session)
    preview = session.get("preview")
    if not isinstance(preview, dict):
        return session
    normalize_denemark_preview(preview)
    normalize_nettomd_preview(preview)
    source_file = str(preview.get("sourceFile") or "")
    current_point = str(preview.get("deliveryPoint") or "")
    current_client = str(preview.get("client") or "")
    file_client, file_delivery_point = recognize_client_from_file_name(source_file)
    if (
        file_client
        and file_delivery_point
        and normalize(file_client) == normalize(current_client)
        and normalize(current_point) in {"", "maindeliverypoint", "unknowndeliverypoint"}
    ):
        preview["deliveryPoint"] = file_delivery_point
        preview["confidence"] = max(int(preview.get("confidence") or 0), 95)
    return session


NETTOMD_DELIVERY_POINTS = {
    "kerpen": "Kerpen",
    "hodenhagen": "Hodenhagen",
    "henstedt": "Henstedt-Ulzburg",
    "henstedtulzburg": "Henstedt-Ulzburg",
    "hamm": "Hamm",
    "ganderkesee": "Ganderkesee",
    "bottrop": "Bottrop",
    "krefeld": "Krefeld",
}


def normalize_nettomd_preview(preview: dict[str, Any]) -> None:
    if normalize(preview.get("client")) != "nettomd" and normalize(preview.get("greenopsFatrans")) != "nettomd":
        return

    current_point = str(preview.get("deliveryPoint") or "")
    customer = str(preview.get("greenopsCustomer") or "")
    point = clean_nettomd_delivery_point(customer) or clean_nettomd_delivery_point(current_point)
    if point and point != current_point:
        preview["client"] = "NettoMD"
        preview["deliveryPoint"] = point
        preview["confidence"] = max(int(preview.get("confidence") or 0), 95)


def clean_nettomd_delivery_point(value: str) -> str:
    normalized = normalize(value)
    for needle, label in NETTOMD_DELIVERY_POINTS.items():
        if needle in normalized:
            return label
    if re.fullmatch(r"(netto)?mdld\d{2}\d{2}", normalized):
        return ""
    return ""


def normalize_denemark_preview(preview: dict[str, Any]) -> None:
    context = " ".join(
        str(preview.get(key) or "")
        for key in [
            "sourceFile",
            "client",
            "deliveryPoint",
            "greenopsCustomer",
            "greenopsFatrans",
        ]
    )
    normalized = context.casefold()
    if not (
        "denemark" in normalized
        or "denmark" in normalized
        or "brabrand" in normalized
        or "køge" in normalized
        or "koge" in normalized
        or "koege" in normalized
    ):
        return

    preview["client"] = "Denemark"
    point_source = str(preview.get("greenopsCustomer") or preview.get("deliveryPoint") or "")
    point = clean_denemark_delivery_point(point_source)
    if point:
        preview["deliveryPoint"] = point
        preview["confidence"] = max(int(preview.get("confidence") or 0), 90)


def clean_denemark_delivery_point(value: str) -> str:
    cleaned = re.sub(r"\s+", " ", str(value or "").strip(" -:\t"))
    cleaned = re.sub(r"^netto\s+", "", cleaned, flags=re.IGNORECASE).strip()
    cleaned = re.sub(r"\bdenem?ark\b", "", cleaned, flags=re.IGNORECASE).strip()
    cleaned = re.sub(r"\b(VCSO|SSCO)\d+\b", "", cleaned, flags=re.IGNORECASE).strip()
    if not cleaned:
        return ""
    normalized = normalize(cleaned)
    if "brabrand" in normalized:
        return "Brabrand"
    if "kge" in normalized or "koge" in normalized or "koege" in normalized:
        return "Køge"
    return cleaned[:80]


def repair_display_texts(value: Any) -> Any:
    if isinstance(value, str):
        return repair_display_text(value)
    if isinstance(value, list):
        return [repair_display_texts(entry) for entry in value]
    if isinstance(value, dict):
        return {key: repair_display_texts(entry) for key, entry in value.items()}
    return value


def repair_display_text(value: str) -> str:
    replacements = {
        "K��ln": "Köln",
        "KÃ¶ln": "Köln",
        "Gro��beeren": "Großbeeren",
        "GroÃŸbeeren": "Großbeeren",
        "K��ge": "Køge",
        "KÃ¸ge": "Køge",
        "G��nzburg": "Günzburg",
        "GÃ¼nzburg": "Günzburg",
        "H??nchen": "Hähnchen",
        "H??hnchen": "Hähnchen",
        "H��nchen": "Hähnchen",
        "H��hnchen": "Hähnchen",
        "Hï¿½ï¿½nchen": "Hähnchen",
        "HÃ¤hnchen": "Hähnchen",
        "K??se": "Käse",
        "K��se": "Käse",
        "Kï¿½ï¿½se": "Käse",
        "KÃ¤se": "Käse",
        "Ziegenk��se": "Ziegenkäse",
        "ZiegenkÃ¤se": "Ziegenkäse",
        "M��ckm��hl": "Möckmühl",
        "MÃ¶ckmÃ¼hl": "Möckmühl",
        "Â·": "-",
    }
    repaired = str(value)
    for source, target in replacements.items():
        repaired = repaired.replace(source, target)
    return repaired


def normalize(value: Any) -> str:
    return re.sub(r"[^a-z0-9]+", "", str(value or "").casefold())


def as_number(value: Any) -> float:
    if value is None:
        return 0.0
    if isinstance(value, (int, float)):
        return float(value)
    text = str(value).strip().replace(" ", "")
    if not text:
        return 0.0
    if "," in text and "." in text:
        text = text.replace(".", "").replace(",", ".")
    else:
        text = text.replace(",", ".")
    try:
        return float(text)
    except ValueError:
        return 0.0


def clean_number(value: float) -> int | float:
    if abs(value - round(value)) < 0.0001:
        return int(round(value))
    return round(value, 2)


def safe_file_part(value: str) -> str:
    cleaned = re.sub(r"[^A-Za-z0-9_-]+", "_", value.strip())
    return cleaned.strip("_") or "export"


def mancos_export_file_name(preview_entries: list[dict[str, Any]], today: datetime | None = None) -> str:
    clients: list[str] = []
    for entry in preview_entries:
        preview = entry.get("preview") if isinstance(entry, dict) else None
        if not isinstance(preview, dict):
            continue
        client = repair_display_text(preview.get("client", "")).strip()
        if client:
            clients.append(client)

    unique_clients = {normalize(client) for client in clients if client}
    if len(unique_clients) == 1 and clients:
        raw_client = clients[0]
    else:
        raw_client = "Mancos"

    client_name = re.sub(r'[<>:"/\\|?*]+', " ", raw_client)
    client_name = re.sub(r"\s+", " ", client_name).strip() or "Mancos"
    export_date = (today or datetime.now()).strftime("%d-%m-%Y")
    return f"{client_name} - {export_date}.xlsx"


def build_day_analytics_workbook(
    sessions: list[dict[str, Any]],
    session_date: str = "",
    havi_uien_settings: dict[str, Any] | None = None,
) -> Workbook:
    filtered_sessions = [
        normalize_saved_session(session)
        for session in sessions
        if isinstance(session, dict) and (not session_date or str(session.get("date") or "") == session_date)
    ]
    if not filtered_sessions:
        raise ValueError("There are no Manco sessions for this day.")

    clients: dict[str, list[dict[str, Any]]] = {}
    for session in filtered_sessions:
        preview = session.get("preview") if isinstance(session.get("preview"), dict) else {}
        client = repair_display_text(str(preview.get("client") or "Unknown client")).strip() or "Unknown client"
        clients.setdefault(client, []).append(session)

    workbook = Workbook()
    workbook.remove(workbook.active)

    dark_green = "0F5132"
    mid_green = "1F7A4D"
    pale_green = "DFF2E6"
    soft_green = "F4FBF6"
    red = "D71920"
    border_color = "B8D8C2"
    thin_border = Border(
        left=Side(style="thin", color=border_color),
        right=Side(style="thin", color=border_color),
        top=Side(style="thin", color=border_color),
        bottom=Side(style="thin", color=border_color),
    )
    headers = ["Delivery point", "Order Reference", "Article", "Description", "Ordered", "Delivered", "Manco", "Manco %"]
    used_sheet_names: set[str] = set()
    used_table_names: set[str] = set()

    for client, client_sessions in sorted(clients.items(), key=lambda item: normalize(item[0])):
        sheet = workbook.create_sheet(unique_sheet_name(client, used_sheet_names))
        sheet.append([f"{client} Manco´s"])
        sheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
        title_cell = sheet.cell(row=1, column=1)
        title_cell.font = Font(bold=True, size=16, color="FFFFFF")
        title_cell.fill = PatternFill("solid", fgColor=dark_green)
        title_cell.alignment = Alignment(horizontal="center", vertical="center")
        sheet.row_dimensions[1].height = 28

        sheet.append(["Date", session_date or "All dates", "Orders", len(client_sessions)])
        for column in range(1, len(headers) + 1):
            cell = sheet.cell(row=2, column=column)
            cell.fill = PatternFill("solid", fgColor=pale_green)
            cell.border = thin_border
            cell.alignment = Alignment(vertical="center")
            if column in {1, 3}:
                cell.font = Font(bold=True, color=dark_green)

        summary_row = 3
        sheet.append(["Total", "", "", "", 0, 0, 0, 0])

        header_row = 5
        for column, header in enumerate(headers, start=1):
            cell = sheet.cell(row=header_row, column=column, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor=mid_green)
            cell.border = thin_border
            cell.alignment = Alignment(horizontal="center", vertical="center")

        current_row = header_row + 1
        ordered_total = delivered_total = shortage_total = 0.0
        for session in client_sessions:
            preview = session.get("preview") if isinstance(session.get("preview"), dict) else {}
            preview = apply_havi_uien_settings_to_preview(preview, havi_uien_settings)
            delivery_point = repair_display_text(str(preview.get("deliveryPoint") or ""))
            reference = repair_display_text(
                str(
                    preview.get("orderReference")
                    or preview.get("greenopsReference")
                    or preview.get("reference")
                    or session.get("reference")
                    or ""
                )
            )
            for item in preview.get("items") or []:
                if not isinstance(item, dict):
                    continue
                ordered = as_number(item.get("orderedQuantity"))
                delivered = as_number(item.get("deliveredQuantity"))
                shortage = as_number(item.get("shortageQuantity"))
                if delivered <= 0 and shortage:
                    delivered = max(ordered - shortage, 0)
                rate = (shortage / ordered * 100) if ordered else 0
                ordered_total += ordered
                delivered_total += delivered
                shortage_total += shortage
                values = [
                    delivery_point,
                    reference,
                    repair_display_text(str(item.get("article") or "")),
                    repair_display_text(str(item.get("description") or "")),
                    clean_number(ordered),
                    clean_number(delivered),
                    clean_number(shortage),
                    round(rate, 2),
                ]
                for column, value in enumerate(values, start=1):
                    cell = sheet.cell(row=current_row, column=column, value=value)
                    cell.border = thin_border
                    cell.fill = PatternFill("solid", fgColor=soft_green if current_row % 2 == 0 else "FFFFFF")
                    if column >= 5:
                        cell.alignment = Alignment(horizontal="right")
                    else:
                        cell.alignment = Alignment(vertical="center", wrap_text=column == 4)
                    if column == 7 and as_number(value) > 0:
                        cell.font = Font(bold=True, color=red)
                current_row += 1

        total_rate = (shortage_total / ordered_total * 100) if ordered_total else 0
        total_values = ["Total", "", "", "", clean_number(ordered_total), clean_number(delivered_total), clean_number(shortage_total), round(total_rate, 2)]
        for column, value in enumerate(total_values, start=1):
            cell = sheet.cell(row=summary_row, column=column, value=value)
            cell.font = Font(bold=True, color=red if column == 7 else dark_green)
            cell.fill = PatternFill("solid", fgColor=pale_green)
            cell.border = thin_border
            if column >= 5:
                cell.alignment = Alignment(horizontal="right")

        last_data_row = max(current_row - 1, header_row)
        if last_data_row > header_row:
            table = Table(displayName=unique_table_name(client, used_table_names), ref=f"A{header_row}:H{last_data_row}")
            table.tableStyleInfo = TableStyleInfo(
                name="TableStyleMedium4",
                showFirstColumn=False,
                showLastColumn=False,
                showRowStripes=True,
                showColumnStripes=False,
            )
            sheet.add_table(table)
        sheet.freeze_panes = "A6"
        widths = {"A": 20, "B": 24, "C": 18, "D": 44, "E": 14, "F": 14, "G": 14, "H": 12}
        for column_letter, width in widths.items():
            sheet.column_dimensions[column_letter].width = width

    return workbook


def apply_havi_uien_settings_to_preview(
    preview: dict[str, Any],
    settings: dict[str, Any] | None,
) -> dict[str, Any]:
    if not settings or not is_havi_uien_preview(preview):
        return preview
    article = str(settings.get("article") or "").strip()
    description = str(settings.get("description") or "").strip()
    if not article and not description:
        return preview
    updated = dict(preview)
    updated["items"] = [
        {
            **item,
            "orderNumber": str(
                item.get("orderNumber")
                or item.get("description")
                or item.get("reference")
                or ""
            ).strip(),
            "article": article,
            "description": description,
            "dc": str(item.get("dc") or " - ".join(
                part for part in [
                    str(item.get("article") or "").strip(),
                    str(item.get("description") or "").strip(),
                ]
                if part
            )).strip(),
        }
        if isinstance(item, dict)
        else item
        for item in preview.get("items") or []
    ]
    return updated


def is_havi_uien_preview(preview: dict[str, Any]) -> bool:
    text = normalize(
        " ".join(
            str(preview.get(key) or "")
            for key in ["client", "deliveryPoint", "greenopsFatrans", "greenopsCustomer"]
        )
    )
    return "havi" in text and ("uien" in text or "onion" in text)


def unique_sheet_name(value: str, used: set[str]) -> str:
    base = re.sub(r"[\[\]\:\*\?\/\\]+", " ", repair_display_text(value)).strip() or "Client"
    base = re.sub(r"\s+", " ", base)[:31].strip() or "Client"
    candidate = base
    index = 2
    while candidate.casefold() in used:
        suffix = f" {index}"
        candidate = f"{base[:31 - len(suffix)]}{suffix}".strip()
        index += 1
    used.add(candidate.casefold())
    return candidate


def unique_table_name(value: str, used: set[str]) -> str:
    base = re.sub(r"[^A-Za-z0-9_]+", "_", repair_display_text(value)).strip("_") or "Client"
    candidate = f"{base[:20]}Table"
    index = 2
    while candidate.casefold() in used:
        candidate = f"{base[:18]}T{index}"
        index += 1
    used.add(candidate.casefold())
    return candidate


def mancos_total_row(rows: list[dict[str, Any]]) -> list[Any]:
    ordered_total = sum(as_number(item.get("orderedQuantity")) for item in rows)
    delivered_total = sum(as_number(item.get("deliveredQuantity")) for item in rows)
    shortage_total = sum(as_number(item.get("shortageQuantity")) for item in rows)
    shortage_rate = (shortage_total / ordered_total * 100) if ordered_total else 0
    return [
        "",
        "",
        "",
        "",
        "Total",
        clean_number(ordered_total),
        clean_number(delivered_total),
        clean_number(shortage_total),
        clean_number(shortage_rate),
    ]


def style_mancos_sheet(sheet, row_count: int) -> None:
    dark_green = "0F5132"
    mid_green = "1F7A4D"
    pale_green = "E2F5E8"
    red = "D71920"

    sheet.merge_cells("A1:I1")
    sheet["A1"].font = Font(bold=True, size=16, color="FFFFFF")
    sheet["A1"].fill = PatternFill("solid", fgColor=dark_green)
    sheet["A1"].alignment = Alignment(horizontal="center")
    sheet.row_dimensions[1].height = 28

    for row in (2, 3):
        sheet[f"A{row}"].font = Font(bold=True, color=dark_green)
        sheet[f"B{row}"].font = Font(bold=True)
        sheet[f"A{row}"].fill = PatternFill("solid", fgColor=pale_green)
        sheet[f"B{row}"].fill = PatternFill("solid", fgColor=pale_green)

    header_row = 5
    for cell in sheet[header_row]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor=mid_green)
        cell.alignment = Alignment(horizontal="center")

    first_data_row = header_row + 1
    last_data_row = header_row + row_count
    if row_count:
        table = Table(displayName="MancosTable", ref=f"A{header_row}:I{last_data_row}")
        table.tableStyleInfo = TableStyleInfo(
            name="TableStyleMedium4",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False,
        )
        sheet.add_table(table)
        for row in sheet.iter_rows(min_row=first_data_row, max_row=last_data_row, min_col=6, max_col=9):
            for cell in row:
                cell.alignment = Alignment(horizontal="right")
        for cell in sheet.iter_rows(min_row=first_data_row, max_row=last_data_row, min_col=8, max_col=8):
            cell[0].font = Font(bold=True, color=red)

    total_row = last_data_row + 1
    if row_count and sheet.max_row >= total_row:
        for cell in sheet[total_row]:
            cell.font = Font(bold=True, color=dark_green)
            cell.fill = PatternFill("solid", fgColor=pale_green)
        for row in sheet.iter_rows(min_row=total_row, max_row=total_row, min_col=6, max_col=9):
            for cell in row:
                cell.alignment = Alignment(horizontal="right")
        sheet[f"H{total_row}"].font = Font(bold=True, color=red)

    sheet.freeze_panes = "A6"
    sheet.auto_filter.ref = f"A{header_row}:I{max(last_data_row, header_row)}"

    min_widths = {
        "A": 18,
        "B": 20,
        "C": 22,
        "D": 18,
        "E": 42,
        "F": 16,
        "G": 18,
        "H": 16,
        "I": 14,
    }
    for column_letter, width in min_widths.items():
        sheet.column_dimensions[column_letter].width = width


def build_mancos_export_rows(preview_entries: list[dict[str, Any]]) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    for entry in preview_entries:
        preview = entry.get("preview") if isinstance(entry, dict) else None
        if not isinstance(preview, dict):
            continue
        client = repair_display_text(preview.get("client", ""))
        delivery_point = repair_display_text(preview.get("deliveryPoint", ""))
        reference = repair_display_text(
            preview.get("greenopsReference")
            or preview.get("reference")
            or entry.get("reference")
            or entry.get("id")
            or ""
        )
        for item in preview.get("items", []):
            if not isinstance(item, dict):
                continue
            ordered = as_number(item.get("orderedQuantity"))
            shortage = as_number(item.get("shortageQuantity"))
            if ordered <= 0 or shortage <= 0:
                continue
            delivered = as_number(item.get("deliveredQuantity"))
            rows.append(
                {
                    "client": client,
                    "deliveryPoint": delivery_point,
                    "orderReference": reference,
                    "article": repair_display_text(item.get("article", "")),
                    "description": repair_display_text(item.get("description", "")),
                    "orderedQuantity": clean_number(ordered),
                    "deliveredQuantity": clean_number(delivered),
                    "shortageQuantity": clean_number(shortage),
                    "shortagePercentage": round((shortage / ordered * 100) if ordered else 0, 2),
                }
            )
    return rows


def build_mancos_export_sections(preview_entries: list[dict[str, Any]]) -> list[dict[str, Any]]:
    sections: list[dict[str, Any]] = []
    for entry in preview_entries:
        preview = entry.get("preview") if isinstance(entry, dict) else None
        if not isinstance(preview, dict):
            continue
        rows: list[dict[str, Any]] = []
        for item in preview.get("items", []):
            if not isinstance(item, dict):
                continue
            ordered = as_number(item.get("orderedQuantity"))
            if ordered <= 0:
                continue
            shortage = as_number(item.get("shortageQuantity"))
            delivered = as_number(item.get("deliveredQuantity"))
            if delivered <= 0 and shortage:
                delivered = max(ordered - shortage, 0)
            rows.append(
                {
                    "article": repair_display_text(item.get("article", "")),
                    "description": repair_display_text(item.get("description", "")),
                    "orderNumber": repair_display_text(
                        item.get("orderNumber")
                        or item.get("reference")
                        or (
                            item.get("dc", "").split(" - ")[-1]
                            if isinstance(item.get("dc"), str) and " - " in item.get("dc", "")
                            else ""
                        )
                    ),
                    "orderedQuantity": clean_number(ordered),
                    "deliveredQuantity": clean_number(delivered),
                    "shortageQuantity": clean_number(shortage),
                    "shortagePercentage": round((shortage / ordered * 100) if ordered else 0, 2),
                }
            )
        if not rows:
            continue
        reference = repair_display_text(
            preview.get("greenopsReference")
            or preview.get("reference")
            or entry.get("reference")
            or entry.get("id")
            or ""
        )
        sections.append(
            {
                "client": repair_display_text(preview.get("client", "")),
                "deliveryPoint": repair_display_text(preview.get("deliveryPoint", "")),
                "orderReference": reference,
                "isHaviUien": is_havi_uien_preview(preview),
                "rows": rows,
            }
        )
    return sections


def build_mancos_export_workbook(preview_entries: list[dict[str, Any]]) -> Workbook:
    sections = build_mancos_export_sections(preview_entries)
    if not sections:
        raise ValueError("There are no order lines to export.")

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Mancos"
    max_columns = 7 if any(section.get("isHaviUien") for section in sections) else 6
    sheet.append(["Manco´s Export"])
    sheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max_columns)
    sheet.row_dimensions[1].height = 28

    dark_green = "0F5132"
    mid_green = "1F7A4D"
    pale_green = "DFF2E6"
    soft_green = "F4FBF6"
    red = "D71920"
    border_color = "B8D8C2"
    thin_border = Border(
        left=Side(style="thin", color=border_color),
        right=Side(style="thin", color=border_color),
        top=Side(style="thin", color=border_color),
        bottom=Side(style="thin", color=border_color),
    )

    sheet["A1"].font = Font(bold=True, size=16, color="FFFFFF")
    sheet["A1"].fill = PatternFill("solid", fgColor=dark_green)
    sheet["A1"].alignment = Alignment(horizontal="center")

    current_row = 3
    for section_index, section in enumerate(sections):
        if section_index:
            current_row += 2

        is_havi_uien_section = bool(section.get("isHaviUien"))
        meta_rows = [
            ("Client", section["client"]),
            ("Delivery point", section["deliveryPoint"]),
        ]
        if not is_havi_uien_section:
            meta_rows.append(("Order Reference", section["orderReference"]))
        for label, value in meta_rows:
            sheet.cell(row=current_row, column=1, value=label)
            sheet.cell(row=current_row, column=2, value=value)
            sheet.merge_cells(start_row=current_row, start_column=2, end_row=current_row, end_column=max_columns)
            for column in range(1, max_columns + 1):
                cell = sheet.cell(row=current_row, column=column)
                cell.fill = PatternFill("solid", fgColor=pale_green)
                cell.border = thin_border
                cell.alignment = Alignment(vertical="center")
            sheet.cell(row=current_row, column=1).font = Font(bold=True, color=dark_green)
            sheet.cell(row=current_row, column=2).font = Font(bold=True)
            current_row += 1

        current_row += 1
        header_row = current_row
        headers = (
            ["Article", "Description", "Order Number", "Ordered", "Delivered", "Manco", "Manco %"]
            if is_havi_uien_section
            else ["Article", "Description", "Ordered", "Delivered", "Manco", "Manco %"]
        )
        numeric_start_column = 4 if is_havi_uien_section else 3
        shortage_column = 6 if is_havi_uien_section else 5
        for column, header in enumerate(headers, start=1):
            cell = sheet.cell(row=header_row, column=column, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor=mid_green)
            cell.border = thin_border
            cell.alignment = Alignment(horizontal="center", vertical="center")
        current_row += 1

        first_item_row = current_row
        ordered_total = delivered_total = shortage_total = 0.0
        for item in section["rows"]:
            ordered = as_number(item.get("orderedQuantity"))
            delivered = as_number(item.get("deliveredQuantity"))
            shortage = as_number(item.get("shortageQuantity"))
            ordered_total += ordered
            delivered_total += delivered
            shortage_total += shortage
            values = [
                item.get("article", ""),
                item.get("description", ""),
                item.get("orderNumber", ""),
                clean_number(ordered),
                clean_number(delivered),
                clean_number(shortage),
                item.get("shortagePercentage", 0),
            ]
            if not is_havi_uien_section:
                values.pop(2)
            for column, value in enumerate(values, start=1):
                cell = sheet.cell(row=current_row, column=column, value=value)
                cell.border = thin_border
                cell.fill = PatternFill("solid", fgColor=soft_green if current_row % 2 == 0 else "FFFFFF")
                if column >= numeric_start_column:
                    cell.alignment = Alignment(horizontal="right")
                if column == shortage_column and as_number(value) > 0:
                    cell.font = Font(bold=True, color=red)
            current_row += 1

        last_item_row = current_row - 1
        last_table_column = "G" if is_havi_uien_section else "F"
        table = Table(displayName=f"MancosOrder{section_index + 1}", ref=f"A{header_row}:{last_table_column}{last_item_row}")
        table.tableStyleInfo = TableStyleInfo(
            name="TableStyleMedium4",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False,
        )
        sheet.add_table(table)

        total_rate = (shortage_total / ordered_total * 100) if ordered_total else 0
        total_values = (
            ["", "Total", "", clean_number(ordered_total), clean_number(delivered_total), clean_number(shortage_total), round(total_rate, 2)]
            if is_havi_uien_section
            else ["", "Total", clean_number(ordered_total), clean_number(delivered_total), clean_number(shortage_total), round(total_rate, 2)]
        )
        for column, value in enumerate(total_values, start=1):
            cell = sheet.cell(row=current_row, column=column, value=value)
            cell.font = Font(bold=True, color=dark_green if column != shortage_column else red)
            cell.fill = PatternFill("solid", fgColor=pale_green)
            cell.border = thin_border
            if column >= numeric_start_column:
                cell.alignment = Alignment(horizontal="right")
        current_row += 1

    sheet.freeze_panes = "A2"
    widths = {"A": 18, "B": 42, "C": 24, "D": 16, "E": 18, "F": 16, "G": 14}
    for column_letter, width in widths.items():
        sheet.column_dimensions[column_letter].width = width
    return workbook


def parse_excel(file_name: str, raw: bytes) -> dict[str, Any]:
    suffix = Path(file_name).suffix.casefold()
    if suffix == ".csv":
        text = raw.decode("utf-8-sig", errors="replace")
        rows = list(csv.reader(StringIO(text)))
    elif suffix in {".xlsx", ".xlsm"}:
        workbook = load_workbook(BytesIO(raw), data_only=True, read_only=True)
        sheet = workbook.active
        rows = [list(row) for row in sheet.iter_rows(values_only=True)]
        workbook.close()
    else:
        raise ValueError("Please upload a .xlsx, .xlsm, or .csv file.")

    rows = [row for row in rows if any(cell not in (None, "") for cell in row)]
    if not rows:
        raise ValueError("The uploaded file is empty.")

    context = " ".join(str(cell) for row in rows[:20] for cell in row if cell not in (None, ""))
    header_index, columns = find_table_header(rows)
    if header_index is None:
        raise ValueError("Could not find columns for article and quantity data.")

    items = extract_items(rows[header_index + 1 :], columns)
    ordered_total = sum(item["orderedQuantity"] for item in items)
    shortage_total = sum(item["shortageQuantity"] for item in items)
    shortage_percentage = (shortage_total / ordered_total * 100) if ordered_total else 0
    recognized = recognize_client(file_name, context)

    return {
        "sourceFile": file_name,
        "client": recognized["client"],
        "deliveryPoint": recognized["deliveryPoint"],
        "confidence": recognized["confidence"],
        "orderedTotal": clean_number(ordered_total),
        "shortageTotal": clean_number(shortage_total),
        "shortagePercentage": round(shortage_percentage, 2),
        "lineCount": len(items),
        "items": items,
        "detectedColumns": columns,
    }


def find_table_header(rows: list[list[Any]]) -> tuple[int | None, dict[str, int]]:
    for row_index, row in enumerate(rows[:80]):
        normalized_headers = [normalize(cell) for cell in row]
        columns = {
            "article": find_column(normalized_headers, ARTICLE_HEADERS),
            "description": find_column(normalized_headers, DESCRIPTION_HEADERS),
            "ordered": find_column(normalized_headers, ORDERED_HEADERS),
            "delivered": find_column(normalized_headers, DELIVERED_HEADERS),
            "shortage": find_column(normalized_headers, SHORTAGE_HEADERS),
        }
        if columns["article"] is not None and columns["ordered"] is not None:
            return row_index, {key: value for key, value in columns.items() if value is not None}
    return None, {}


def find_column(headers: list[str], candidates: list[str]) -> int | None:
    normalized_candidates = [normalize(candidate) for candidate in candidates]
    for index, header in enumerate(headers):
        if not header:
            continue
        if header in normalized_candidates:
            return index
        if any(candidate in header for candidate in normalized_candidates):
            return index
    return None


def extract_items(rows: list[list[Any]], columns: dict[str, int]) -> list[dict[str, Any]]:
    items: list[dict[str, Any]] = []
    for row in rows:
        article = get_cell(row, columns["article"])
        if article in (None, ""):
            continue
        ordered = as_number(get_cell(row, columns["ordered"]))
        if ordered <= 0:
            continue
        shortage = 0.0
        delivered = ordered
        shortage_percentage = 0.0
        items.append(
            {
                "article": str(article).strip(),
                "description": str(get_cell(row, columns.get("description")) or "").strip(),
                "orderedQuantity": clean_number(ordered),
                "deliveredQuantity": clean_number(delivered),
                "shortageQuantity": clean_number(shortage),
                "shortagePercentage": round(shortage_percentage, 2),
            }
        )
    return items


def get_cell(row: list[Any], index: int | None) -> Any:
    if index is None or index >= len(row):
        return None
    return row[index]


def recognize_client(file_name: str, context: str = "") -> dict[str, str | int]:
    file_client, file_delivery_point = recognize_client_from_file_name(file_name)
    text = f"{file_name} {context}"
    normalized_text = normalize(text)
    best_client = "Unknown client"
    best_point = "Unknown delivery point"
    best_score = 0

    for profile in CLIENT_PROFILES:
        client_score = 0
        for alias in profile["aliases"]:
            if normalize(alias) in normalized_text:
                client_score = max(client_score, 50)
        point_name = "Main delivery point"
        point_score = 0
        for name, aliases in profile["delivery_points"].items():
            for alias in aliases:
                if normalize(alias) in normalized_text:
                    point_name = name
                    point_score = max(point_score, 45)
        score = client_score + point_score
        if score > best_score:
            best_score = score
            best_client = str(profile["client"])
            best_point = point_name if point_score else "Main delivery point"

    if file_client and (best_score == 0 or normalize(file_client) == normalize(best_client)):
        return {
            "client": file_client,
            "deliveryPoint": file_delivery_point or best_point,
            "confidence": 95 if file_delivery_point else max(best_score, 50),
        }

    if best_score == 0:
        return {"client": best_client, "deliveryPoint": best_point, "confidence": 0}
    return {"client": best_client, "deliveryPoint": best_point, "confidence": min(best_score, 95)}


def recognize_client_from_file_name(file_name: str) -> tuple[str | None, str | None]:
    stem = Path(file_name).stem.strip()
    parts = [part.strip() for part in re.split(r"[_\-\s]+", stem) if part.strip()]
    if not parts:
        return None, None

    first_part = normalize(parts[0])
    for profile in CLIENT_PROFILES:
        for alias in profile["aliases"]:
            if first_part == normalize(alias):
                delivery_point = " ".join(parts[1:]).strip() or None
                return str(profile["client"]), delivery_point
    return None, None


def multipart_file(content_type: str, body: bytes) -> tuple[str, bytes]:
    match = re.search(r"boundary=(.+)", content_type)
    if not match:
        raise ValueError("Invalid upload request.")
    boundary = match.group(1).strip('"').encode()
    for part in body.split(b"--" + boundary):
        if b"Content-Disposition:" not in part:
            continue
        header_blob, _, data = part.partition(b"\r\n\r\n")
        if not data:
            continue
        headers = header_blob.decode("utf-8", errors="replace")
        name_match = re.search(r'filename="([^"]+)"', headers)
        if not name_match:
            continue
        file_name = Path(name_match.group(1)).name
        return file_name, data.rstrip(b"\r\n-")
    raise ValueError("No file was uploaded.")


class ShortageHandler(BaseHTTPRequestHandler):
    def do_GET(self) -> None:
        if self.path in {"/", "/index.html"}:
            self.serve_file(STATIC_DIR / "index.html", "text/html; charset=utf-8")
            return
        if self.path == "/styles.css":
            self.serve_file(STATIC_DIR / "styles.css", "text/css; charset=utf-8")
            return
        if self.path == "/app.js":
            self.serve_file(STATIC_DIR / "app.js", "application/javascript; charset=utf-8")
            return
        if self.path.startswith("/api/day-sessions"):
            self.json_response({"daySessions": load_day_sessions()})
            return
        if self.path.startswith("/api/sessions"):
            self.json_response({"sessions": load_sessions()})
            return
        self.send_error(HTTPStatus.NOT_FOUND, "Not found")

    def do_POST(self) -> None:
        if self.path == "/api/shortages/parse":
            self.handle_parse()
            return
        if self.path == "/api/export-mancos":
            self.handle_export_mancos()
            return
        if self.path == "/api/export-day-analytics":
            self.handle_export_day_analytics()
            return
        if self.path == "/api/day-sessions":
            self.handle_save_day_session()
            return
        if self.path == "/api/sessions":
            self.handle_save_session()
            return
        self.send_error(HTTPStatus.NOT_FOUND, "Not found")

    def do_PUT(self) -> None:
        if self.path == "/api/sessions":
            self.handle_update_session()
            return
        self.send_error(HTTPStatus.NOT_FOUND, "Not found")

    def do_DELETE(self) -> None:
        if self.path == "/api/day-sessions":
            self.handle_delete_day_session()
            return
        self.send_error(HTTPStatus.NOT_FOUND, "Not found")

    def handle_parse(self) -> None:
        length = int(self.headers.get("content-length", "0"))
        if length <= 0 or length > 12_000_000:
            self.json_response({"error": "Invalid upload size."}, HTTPStatus.BAD_REQUEST)
            return
        try:
            file_name, raw = multipart_file(self.headers.get("content-type", ""), self.rfile.read(length))
            result = parse_excel(file_name, raw)
        except ValueError as exc:
            self.json_response({"error": str(exc)}, HTTPStatus.BAD_REQUEST)
            return
        except Exception as exc:
            self.json_response({"error": f"Could not read the file: {exc}"}, HTTPStatus.BAD_REQUEST)
            return
        self.json_response({"preview": result})

    def handle_export_mancos(self) -> None:
        length = int(self.headers.get("content-length", "0"))
        if length <= 0 or length > 8_000_000:
            self.json_response({"error": "Invalid export payload."}, HTTPStatus.BAD_REQUEST)
            return
        try:
            payload = json.loads(self.rfile.read(length).decode("utf-8"))
        except (UnicodeDecodeError, json.JSONDecodeError):
            self.json_response({"error": "Invalid JSON body."}, HTTPStatus.BAD_REQUEST)
            return

        if isinstance(payload.get("previews"), list):
            preview_entries = [entry for entry in payload.get("previews", []) if isinstance(entry, dict)]
        elif isinstance(payload.get("preview"), dict):
            preview_entries = [{"preview": payload["preview"]}]
        else:
            self.json_response({"error": "Missing delivery point data."}, HTTPStatus.BAD_REQUEST)
            return

        try:
            workbook = build_mancos_export_workbook(preview_entries)
        except ValueError as exc:
            self.json_response({"error": str(exc)}, HTTPStatus.BAD_REQUEST)
            return

        output = BytesIO()
        workbook.save(output)
        data = output.getvalue()
        file_name = mancos_export_file_name(preview_entries)
        self.send_response(HTTPStatus.OK)
        self.send_header("Content-Type", "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        self.send_header("Content-Disposition", f'attachment; filename="{file_name}"')
        self.send_header("Content-Length", str(len(data)))
        self.end_headers()
        self.wfile.write(data)

    def handle_export_day_analytics(self) -> None:
        length = int(self.headers.get("content-length", "0"))
        if length <= 0 or length > 400_000:
            self.json_response({"error": "Invalid export payload."}, HTTPStatus.BAD_REQUEST)
            return
        try:
            payload = json.loads(self.rfile.read(length).decode("utf-8") or "{}")
        except (UnicodeDecodeError, json.JSONDecodeError):
            self.json_response({"error": "Invalid JSON body."}, HTTPStatus.BAD_REQUEST)
            return

        session_date = str(payload.get("date") or "").strip()
        work_session_id = str(payload.get("workSessionId") or "").strip()
        havi_uien_settings = payload.get("haviUienSettings") if isinstance(payload.get("haviUienSettings"), dict) else None
        sessions = load_sessions()
        if work_session_id:
            sessions = [session for session in sessions if str(session.get("workSessionId") or "") == work_session_id]
        try:
            workbook = build_day_analytics_workbook(sessions, session_date, havi_uien_settings)
        except ValueError as exc:
            self.json_response({"error": str(exc)}, HTTPStatus.BAD_REQUEST)
            return

        output = BytesIO()
        workbook.save(output)
        data = output.getvalue()
        safe_date = session_date or datetime.now().strftime("%Y-%m-%d")
        file_name = f"Manco Analytics - {safe_date}.xlsx"
        self.send_response(HTTPStatus.OK)
        self.send_header("Content-Type", "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        self.send_header("Content-Disposition", f'attachment; filename="{file_name}"')
        self.send_header("Content-Length", str(len(data)))
        self.end_headers()
        self.wfile.write(data)

    def handle_save_session(self) -> None:
        length = int(self.headers.get("content-length", "0"))
        if length <= 0 or length > 2_000_000:
            self.json_response({"error": "Invalid session payload."}, HTTPStatus.BAD_REQUEST)
            return
        try:
            payload = json.loads(self.rfile.read(length).decode("utf-8"))
        except (UnicodeDecodeError, json.JSONDecodeError):
            self.json_response({"error": "Invalid JSON body."}, HTTPStatus.BAD_REQUEST)
            return
        session_date = str(payload.get("date", "")).strip()
        try:
            parsed_date = date.fromisoformat(session_date)
        except ValueError:
            self.json_response({"error": "Please choose a valid session date."}, HTTPStatus.BAD_REQUEST)
            return
        preview = payload.get("preview")
        if not isinstance(preview, dict):
            self.json_response({"error": "Upload an Excel file before saving the session."}, HTTPStatus.BAD_REQUEST)
            return

        sessions = load_sessions()
        session = {
            "id": str(uuid.uuid4()),
            "date": session_date,
            "weekday": parsed_date.strftime("%A"),
            "name": str(payload.get("name") or f"{parsed_date.strftime('%A')} Manco session"),
            "workSessionId": str(payload.get("workSessionId") or "").strip(),
            "createdAt": datetime.now().isoformat(timespec="seconds"),
            "preview": preview,
        }
        sessions.insert(0, session)
        save_sessions(sessions[:1000])
        self.json_response({"session": session, "sessions": sessions[:1000]})

    def handle_save_day_session(self) -> None:
        length = int(self.headers.get("content-length", "0"))
        if length <= 0 or length > 200_000:
            self.json_response({"error": "Invalid session payload."}, HTTPStatus.BAD_REQUEST)
            return
        try:
            payload = json.loads(self.rfile.read(length).decode("utf-8"))
        except (UnicodeDecodeError, json.JSONDecodeError):
            self.json_response({"error": "Invalid JSON body."}, HTTPStatus.BAD_REQUEST)
            return

        session_date = str(payload.get("date", "")).strip()
        try:
            parsed_date = date.fromisoformat(session_date)
        except ValueError:
            self.json_response({"error": "Please choose a valid session date."}, HTTPStatus.BAD_REQUEST)
            return

        sessions = load_day_sessions()
        existing = next((entry for entry in sessions if entry.get("date") == session_date), None)
        if existing:
            self.json_response({"daySession": existing, "daySessions": sessions})
            return

        day_session = {
            "id": str(uuid.uuid4()),
            "date": session_date,
            "weekday": parsed_date.strftime("%A"),
            "name": str(payload.get("name") or f"{parsed_date.strftime('%A')} Manco review"),
            "createdAt": datetime.now().isoformat(timespec="seconds"),
        }
        sessions.insert(0, day_session)
        sessions = sorted(sessions, key=lambda entry: str(entry.get("date", "")), reverse=True)[:1000]
        save_day_sessions(sessions)
        self.json_response({"daySession": day_session, "daySessions": sessions})

    def handle_delete_day_session(self) -> None:
        length = int(self.headers.get("content-length", "0"))
        if length <= 0 or length > 200_000:
            self.json_response({"error": "Invalid delete payload."}, HTTPStatus.BAD_REQUEST)
            return
        try:
            payload = json.loads(self.rfile.read(length).decode("utf-8"))
        except (UnicodeDecodeError, json.JSONDecodeError):
            self.json_response({"error": "Invalid JSON body."}, HTTPStatus.BAD_REQUEST)
            return

        session_id = str(payload.get("id") or "").strip()
        session_date = str(payload.get("date") or "").strip()
        if not session_id and not session_date:
            self.json_response({"error": "Missing session id or date."}, HTTPStatus.BAD_REQUEST)
            return

        day_sessions = load_day_sessions()
        matched_date = session_date
        for day_session in day_sessions:
            if session_id and day_session.get("id") == session_id:
                matched_date = str(day_session.get("date") or "")
                break

        if not matched_date:
            self.json_response({"error": "Session date not found."}, HTTPStatus.NOT_FOUND)
            return

        day_sessions = [
            entry
            for entry in day_sessions
            if entry.get("id") != session_id and entry.get("date") != matched_date
        ]
        order_sessions = [entry for entry in load_sessions() if entry.get("date") != matched_date]
        save_day_sessions(day_sessions)
        save_sessions(order_sessions[:1000])
        self.json_response({"ok": True, "daySessions": day_sessions, "sessions": order_sessions[:1000]})

    def handle_update_session(self) -> None:
        length = int(self.headers.get("content-length", "0"))
        if length <= 0 or length > 2_000_000:
            self.json_response({"error": "Invalid session payload."}, HTTPStatus.BAD_REQUEST)
            return
        try:
            payload = json.loads(self.rfile.read(length).decode("utf-8"))
        except (UnicodeDecodeError, json.JSONDecodeError):
            self.json_response({"error": "Invalid JSON body."}, HTTPStatus.BAD_REQUEST)
            return

        session_id = str(payload.get("id", "")).strip()
        preview = payload.get("preview")
        if not session_id or not isinstance(preview, dict):
            self.json_response({"error": "Missing session or preview data."}, HTTPStatus.BAD_REQUEST)
            return

        sessions = load_sessions()
        for session in sessions:
            if session.get("id") == session_id:
                session["preview"] = preview
                if payload.get("workSessionId"):
                    session["workSessionId"] = str(payload.get("workSessionId") or "").strip()
                session["updatedAt"] = datetime.now().isoformat(timespec="seconds")
                save_sessions(sessions[:1000])
                self.json_response({"session": session, "sessions": sessions[:1000]})
                return

        self.json_response({"error": "Session not found."}, HTTPStatus.NOT_FOUND)

    def serve_file(self, path: Path, content_type: str) -> None:
        if not path.is_file():
            self.send_error(HTTPStatus.NOT_FOUND, "Not found")
            return
        data = path.read_bytes()
        self.send_response(HTTPStatus.OK)
        self.send_header("Content-Type", content_type)
        self.send_header("Content-Length", str(len(data)))
        self.end_headers()
        self.wfile.write(data)

    def json_response(self, payload: dict[str, Any], status: HTTPStatus = HTTPStatus.OK) -> None:
        data = json.dumps(payload, ensure_ascii=False).encode("utf-8")
        self.send_response(status)
        self.send_header("Content-Type", "application/json; charset=utf-8")
        self.send_header("Content-Length", str(len(data)))
        self.end_headers()
        self.wfile.write(data)


def main() -> None:
    server = ThreadingHTTPServer((HOST, PORT), ShortageHandler)
    print(f"Manco Tracker running at http://{HOST}:{PORT}")
    server.serve_forever()


if __name__ == "__main__":
    main()
